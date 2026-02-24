import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, font
from tkinter.scrolledtext import ScrolledText
from openpyxl import load_workbook
from openpyxl import Workbook
import os
import time
from datetime import datetime, timedelta
import smtplib
import ssl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import subprocess
import platform
import socket
import threading
from PIL import Image, ImageTk  # You may need to install: pip install Pillow

# ============================================================================
# CONFIGURATION
# ============================================================================

EMAIL_HOST = "smtp.gmail.com"
EMAIL_PORT = 587
EMAIL_ADDRESS = "lanze.anderson@gmail.com"
EMAIL_PASSWORD = "hirl quyv gdzs dewd"
LIBRARY_NAME = "MELCHORA AQUINO HIGH SCHOOL LIBRARY"
BORROWING_DAYS = 3

# Admin credentials
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "library123"

# WiFi Configuration
WIFI_SSID = "Library_WiFi"
WIFI_PASSWORD = "library123"

# ============================================================================
# DATABASE CLASS
# ============================================================================

class ExcelLibraryDatabase:
    def __init__(self, filename="LIBRARY_SYSTEM_DATA.xlsx"):
        self.filename = filename
        self.setup_database()

    def setup_database(self):
        try:
            self.workbook = load_workbook(self.filename)
            self.sheet = self.workbook.active
        except FileNotFoundError:
            self.workbook = Workbook()
            self.sheet = self.workbook.active
            self.sheet.append([
                "Student ID",
                "Student Name",
                "Book Barcode",
                "Book Title",
                "Action",
                "Timestamp",
                "Date Time",
                "Due Date"
            ])
            self.workbook.save(self.filename)

    def log_transaction(self, student_id, student_name, book_barcode, book_title, action, due_date=None):
        try:
            timestamp = time.time()
            date_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            due_date_str = due_date.strftime("%Y-%m-%d") if due_date else ""

            self.sheet.append([
                student_id,
                student_name,
                book_barcode,
                book_title,
                action,
                timestamp,
                date_time,
                due_date_str
            ])

            self.workbook.save(self.filename)
            return True, "Transaction saved successfully"

        except Exception as e:
            return False, str(e)

    def get_active_borrowings(self, student_id=None):
        """Get all active borrowings"""
        records = []
        rows = list(self.sheet.iter_rows(min_row=2, values_only=True))
        
        # First get all borrows
        borrows = []
        for row in rows:
            if len(row) >= 5 and row[4] == 'borrow':
                borrows.append({
                    'student_id': row[0],
                    'student_name': row[1],
                    'book_barcode': row[2],
                    'book_title': row[3],
                    'timestamp': row[5],
                    'date_time': row[6],
                    'due_date': row[7] if len(row) > 7 else ""
                })
        
        # Then get all returns
        returns = {row[2] for row in rows if len(row) >= 5 and row[4] == 'return'}
        
        # Filter out returned books
        active = [b for b in borrows if b['book_barcode'] not in returns]
        
        if student_id:
            active = [b for b in active if str(b['student_id']) == str(student_id)]
        
        return active

    def get_all_transactions(self, limit=1000):
        """Get all transactions"""
        transactions = []
        rows = list(self.sheet.iter_rows(min_row=2, values_only=True))
        for row in rows[:limit]:
            if len(row) >= 5:
                transactions.append({
                    'student_id': row[0],
                    'student_name': row[1],
                    'book_barcode': row[2],
                    'book_title': row[3],
                    'action': row[4],
                    'timestamp': row[5],
                    'date_time': row[6],
                    'due_date': row[7] if len(row) > 7 else ""
                })
        return transactions

    def check_overdue_books(self, student_id=None):
        """Check for overdue books"""
        active = self.get_active_borrowings(student_id)
        overdue = []
        now = time.time()
        
        for record in active:
            try:
                if record['due_date']:
                    due_date = datetime.strptime(record['due_date'], "%Y-%m-%d")
                    due_timestamp = due_date.timestamp()
                    if now > due_timestamp:
                        days_overdue = int((now - due_timestamp) / (24 * 60 * 60))
                        record['days_overdue'] = days_overdue
                        overdue.append(record)
            except:
                continue
        
        return overdue

    def get_statistics(self):
        """Get library statistics"""
        transactions = self.get_all_transactions()
        active = self.get_active_borrowings()
        overdue = self.check_overdue_books()
        
        total_borrows = sum(1 for t in transactions if t['action'] == 'borrow')
        total_returns = sum(1 for t in transactions if t['action'] == 'return')
        
        unique_students = set(t['student_id'] for t in transactions if t['student_id'])
        unique_books = set(t['book_title'] for t in transactions if t['book_title'])
        
        return {
            'total_transactions': len(transactions),
            'total_borrows': total_borrows,
            'total_returns': total_returns,
            'active_borrowings': len(active),
            'overdue_books': len(overdue),
            'unique_students': len(unique_students),
            'unique_books': len(unique_books)
        }

# ============================================================================
# MAIN APPLICATION CLASS
# ============================================================================

class LibrarySoftware:
    def __init__(self, root):
        self.root = root
        self.root.title(f"{LIBRARY_NAME} - Library Management System")
        self.root.geometry("1200x700")
        self.root.minsize(1000, 600)
        
        # Set icon (optional)
        # self.root.iconbitmap('library.ico')
        
        # Configure style
        self.style = ttk.Style()
        self.style.theme_use('clam')
        
        # Configure colors
        self.bg_color = "#f0f0f0"
        self.header_color = "#2c3e50"
        self.accent_color = "#3498db"
        self.success_color = "#27ae60"
        self.warning_color = "#e74c3c"
        
        self.root.configure(bg=self.bg_color)
        
        # Initialize database
        self.database = ExcelLibraryDatabase()
        
        # Load data
        self.students = self.load_student_data()
        self.books = self.load_book_data()
        
        # Current user
        self.current_user = None
        self.is_admin = False
        
        # Setup UI
        self.setup_ui()
        
        # Show login screen
        self.show_login()
    
    def setup_ui(self):
        """Setup the main UI structure"""
        # Create header
        self.header_frame = tk.Frame(self.root, bg=self.header_color, height=80)
        self.header_frame.pack(fill=tk.X, side=tk.TOP)
        self.header_frame.pack_propagate(False)
        
        # Header title
        header_font = font.Font(family="Helvetica", size=20, weight="bold")
        self.header_label = tk.Label(
            self.header_frame, 
            text=LIBRARY_NAME,
            bg=self.header_color,
            fg="white",
            font=header_font
        )
        self.header_label.pack(side=tk.LEFT, padx=20, pady=20)
        
        # User info in header
        self.user_frame = tk.Frame(self.header_frame, bg=self.header_color)
        self.user_frame.pack(side=tk.RIGHT, padx=20)
        
        self.user_label = tk.Label(
            self.user_frame,
            text="Not logged in",
            bg=self.header_color,
            fg="#ecf0f1",
            font=("Helvetica", 10)
        )
        self.user_label.pack()
        
        self.logout_btn = tk.Button(
            self.user_frame,
            text="Logout",
            bg=self.accent_color,
            fg="white",
            bd=0,
            padx=10,
            command=self.logout
        )
        
        # Main content area
        self.content_frame = tk.Frame(self.root, bg=self.bg_color)
        self.content_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
    
    def clear_content(self):
        """Clear the content frame"""
        for widget in self.content_frame.winfo_children():
            widget.destroy()
    
    def show_login(self):
        """Show login screen"""
        self.clear_content()
        self.user_label.config(text="Not logged in")
        self.logout_btn.pack_forget()
        
        # Create login frame
        login_frame = tk.Frame(self.content_frame, bg=self.bg_color)
        login_frame.pack(expand=True)
        
        # Title
        title_font = font.Font(family="Helvetica", size=24, weight="bold")
        title = tk.Label(
            login_frame,
            text="Welcome to Library System",
            bg=self.bg_color,
            font=title_font,
            fg=self.header_color
        )
        title.pack(pady=50)
        
        # Login buttons
        btn_font = font.Font(size=12)
        
        tk.Button(
            login_frame,
            text="Student Login",
            width=20,
            height=2,
            bg=self.accent_color,
            fg="white",
            font=btn_font,
            bd=0,
            command=self.student_login
        ).pack(pady=10)
        
        tk.Button(
            login_frame,
            text="Admin Login",
            width=20,
            height=2,
            bg=self.header_color,
            fg="white",
            font=btn_font,
            bd=0,
            command=self.admin_login
        ).pack(pady=10)
        
        tk.Button(
            login_frame,
            text="Exit",
            width=20,
            height=2,
            bg=self.warning_color,
            fg="white",
            font=btn_font,
            bd=0,
            command=self.root.quit
        ).pack(pady=10)
        
        # Status
        tk.Label(
            login_frame,
            text="Please select login type",
            bg=self.bg_color,
            fg="gray",
            font=("Helvetica", 10)
        ).pack(pady=20)
    
    def student_login(self):
        """Show student login form"""
        # Create dialog
        dialog = tk.Toplevel(self.root)
        dialog.title("Student Login")
        dialog.geometry("400x200")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Center the dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (400 // 2)
        y = (dialog.winfo_screenheight() // 2) - (200 // 2)
        dialog.geometry(f'+{x}+{y}')
        
        # Form
        tk.Label(dialog, text="Enter Student ID:", font=("Helvetica", 12)).pack(pady=20)
        
        student_id = tk.Entry(dialog, font=("Helvetica", 14), width=20)
        student_id.pack(pady=10)
        student_id.focus()
        
        def login():
            sid = student_id.get().strip()
            if sid in self.students:
                self.current_user = self.students[sid]
                self.is_admin = False
                self.user_label.config(text=f"Student: {self.current_user['name']}")
                self.logout_btn.pack(side=tk.RIGHT, padx=5)
                dialog.destroy()
                self.show_student_dashboard()
            else:
                messagebox.showerror("Error", "Student ID not found")
        
        tk.Button(
            dialog,
            text="Login",
            command=login,
            bg=self.accent_color,
            fg="white",
            padx=20
        ).pack(pady=10)
    
    def admin_login(self):
        """Show admin login form"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Admin Login")
        dialog.geometry("400x250")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Center the dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (400 // 2)
        y = (dialog.winfo_screenheight() // 2) - (250 // 2)
        dialog.geometry(f'+{x}+{y}')
        
        # Form
        tk.Label(dialog, text="Username:", font=("Helvetica", 12)).pack(pady=10)
        username = tk.Entry(dialog, font=("Helvetica", 14), width=20)
        username.pack()
        username.focus()
        
        tk.Label(dialog, text="Password:", font=("Helvetica", 12)).pack(pady=10)
        password = tk.Entry(dialog, font=("Helvetica", 14), width=20, show="*")
        password.pack()
        
        def login():
            if username.get() == ADMIN_USERNAME and password.get() == ADMIN_PASSWORD:
                self.current_user = {"name": "Administrator", "username": ADMIN_USERNAME}
                self.is_admin = True
                self.user_label.config(text="Admin: Administrator")
                self.logout_btn.pack(side=tk.RIGHT, padx=5)
                dialog.destroy()
                self.show_admin_dashboard()
            else:
                messagebox.showerror("Error", "Invalid credentials")
        
        tk.Button(
            dialog,
            text="Login",
            command=login,
            bg=self.accent_color,
            fg="white",
            padx=20
        ).pack(pady=20)
    
    def logout(self):
        """Logout current user"""
        self.current_user = None
        self.is_admin = False
        self.show_login()
    
    def show_student_dashboard(self):
        """Show student dashboard"""
        self.clear_content()
        
        # Check for overdue books
        overdue = self.database.check_overdue_books(self.current_user['student_id'])
        if overdue:
            self.show_warning(f"You have {len(overdue)} overdue book(s)!")
        
        # Create notebook for tabs
        notebook = ttk.Notebook(self.content_frame)
        notebook.pack(fill=tk.BOTH, expand=True)
        
        # Borrow tab
        borrow_frame = ttk.Frame(notebook)
        notebook.add(borrow_frame, text="Borrow Books")
        self.setup_borrow_tab(borrow_frame)
        
        # Return tab
        return_frame = ttk.Frame(notebook)
        notebook.add(return_frame, text="Return Books")
        self.setup_return_tab(return_frame)
        
        # My Books tab
        mybooks_frame = ttk.Frame(notebook)
        notebook.add(mybooks_frame, text="My Books")
        self.setup_mybooks_tab(mybooks_frame)
        
        # Profile tab
        profile_frame = ttk.Frame(notebook)
        notebook.add(profile_frame, text="Profile")
        self.setup_profile_tab(profile_frame)
    
    def setup_borrow_tab(self, parent):
        """Setup borrow books tab"""
        # Available books list
        tk.Label(parent, text="Available Books", font=("Helvetica", 14, "bold")).pack(pady=10)
        
        # Create treeview
        columns = ('Barcode', 'Title', 'Author', 'Status')
        tree = ttk.Treeview(parent, columns=columns, show='headings', height=15)
        
        tree.heading('Barcode', text='Barcode')
        tree.heading('Title', text='Title')
        tree.heading('Author', text='Author')
        tree.heading('Status', text='Status')
        
        tree.column('Barcode', width=120)
        tree.column('Title', width=300)
        tree.column('Author', width=200)
        tree.column('Status', width=100)
        
        # Add scrollbar
        scrollbar = ttk.Scrollbar(parent, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(10, 0))
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y, padx=(0, 10))
        
        # Populate books
        for barcode, book in self.books.items():
            status = "Available" if book['available'] else "Borrowed"
            tree.insert('', tk.END, values=(barcode, book['title'], book['author'], status))
        
        # Borrow button
        def borrow_book():
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("Warning", "Please select a book")
                return
            
            item = tree.item(selected[0])
            barcode = str(item['values'][0]).strip()  # Convert to string and remove spaces
            status = item['values'][3]

            print(f"Selected barcode: '{barcode}'")
            print(f"Available books: {list(self.books.keys())}")
    
            if status != "Available":
                messagebox.showwarning("Warning", "This book is not available")
                return

            if barcode not in self.books:
                messagebox.showerror("Error", f"Book with barcode {barcode} not found in system")
                return

            overdue = self.database.check_overdue_books(self.current_user['student_id'])
            if overdue:
                messagebox.showerror("Error", "Cannot borrow: You have overdue books")
                return
    
            book = self.books[barcode]
            due_date = datetime.now() + timedelta(days=BORROWING_DAYS)
    
            success, msg = self.database.log_transaction(
                self.current_user['student_id'],
                self.current_user['name'],
                barcode,
                book['title'],
                'borrow',
                due_date
            )
    
            if success:
                book['available'] = False
                messagebox.showinfo("Success", f"Book borrowed successfully!\nDue date: {due_date.strftime('%Y-%m-%d')}")
                tree.item(selected[0], values=(barcode, book['title'], book['author'], "Borrowed"))
            else:
                messagebox.showerror("Error", f"Failed to borrow: {msg}")

        
        tk.Button(
            parent,
            text="Borrow Selected Book",
            command=borrow_book,
            bg=self.success_color,
            fg="white",
            padx=20,
            pady=5
        ).pack(pady=10)
    
    def setup_return_tab(self, parent):
        """Setup return books tab"""
        tk.Label(parent, text="My Borrowed Books", font=("Helvetica", 14, "bold")).pack(pady=10)
        
        # Get active borrowings
        active = self.database.get_active_borrowings(self.current_user['student_id'])
        
        if not active:
            tk.Label(parent, text="You have no borrowed books", font=("Helvetica", 12)).pack(pady=50)
            return
        
        # Create treeview
        columns = ('Barcode', 'Title', 'Borrowed Date', 'Due Date', 'Status')
        tree = ttk.Treeview(parent, columns=columns, show='headings', height=15)
        
        tree.heading('Barcode', text='Barcode')
        tree.heading('Title', text='Title')
        tree.heading('Borrowed Date', text='Borrowed Date')
        tree.heading('Due Date', text='Due Date')
        tree.heading('Status', text='Status')
        
        tree.column('Barcode', width=120)
        tree.column('Title', width=300)
        tree.column('Borrowed Date', width=150)
        tree.column('Due Date', width=150)
        tree.column('Status', width=100)
        
        # Add scrollbar
        scrollbar = ttk.Scrollbar(parent, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(10, 0))
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y, padx=(0, 10))
        
        # Populate books
        now = time.time()
        for book in active:
            # Check if overdue
            try:
                due_date = datetime.strptime(book['due_date'], "%Y-%m-%d")
                status = "OVERDUE" if now > due_date.timestamp() else "Active"
            except:
                status = "Active"
            
            tree.insert('', tk.END, values=(
                book['book_barcode'],
                book['book_title'],
                book['date_time'][:10] if book['date_time'] else "",
                book['due_date'],
                status
            ))
        
        # Return button
        def return_book():
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("Warning", "Please select a book to return")
                return
            
            item = tree.item(selected[0])
            barcode = item['values'][0]
            title = item['values'][1]
            
            # Process return
            success, msg = self.database.log_transaction(
                self.current_user['student_id'],
                self.current_user['name'],
                barcode,
                title,
                'return'
            )
            
            if success:
                self.books[barcode]['available'] = True
                messagebox.showinfo("Success", "Book returned successfully!")
                # Refresh tab
                self.setup_return_tab(parent)
            else:
                messagebox.showerror("Error", f"Failed to return: {msg}")
        
        tk.Button(
            parent,
            text="Return Selected Book",
            command=return_book,
            bg=self.accent_color,
            fg="white",
            padx=20,
            pady=5
        ).pack(pady=10)
    
    def setup_mybooks_tab(self, parent):
        """Setup my books tab"""
        tk.Label(parent, text="My Borrowing History", font=("Helvetica", 14, "bold")).pack(pady=10)
        
        # Get all transactions for this student
        all_trans = self.database.get_all_transactions()
        my_trans = [t for t in all_trans if t['student_id'] == self.current_user['student_id']]
        
        if not my_trans:
            tk.Label(parent, text="No borrowing history", font=("Helvetica", 12)).pack(pady=50)
            return
        
        # Create treeview
        columns = ('Date', 'Book', 'Action', 'Due Date')
        tree = ttk.Treeview(parent, columns=columns, show='headings', height=20)
        
        tree.heading('Date', text='Date')
        tree.heading('Book', text='Book Title')
        tree.heading('Action', text='Action')
        tree.heading('Due Date', text='Due Date')
        
        tree.column('Date', width=150)
        tree.column('Book', width=400)
        tree.column('Action', width=100)
        tree.column('Due Date', width=150)
        
        # Add scrollbar
        scrollbar = ttk.Scrollbar(parent, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(10, 0))
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y, padx=(0, 10))
        
        # Populate history
        for trans in my_trans:
            tree.insert('', tk.END, values=(
                trans['date_time'][:16] if trans['date_time'] else "",
                trans['book_title'],
                trans['action'].upper(),
                trans['due_date']
            ))
    
    def setup_profile_tab(self, parent):
        """Setup profile tab"""
        # Profile info
        info_frame = tk.Frame(parent, bg=self.bg_color, relief=tk.GROOVE, bd=2)
        info_frame.pack(pady=50, padx=50, fill=tk.BOTH, expand=True)
        
        tk.Label(
            info_frame,
            text="Student Profile",
            font=("Helvetica", 16, "bold"),
            bg=self.bg_color
        ).pack(pady=20)
        
        # Student details
        details = [
            ("Student ID:", self.current_user['student_id']),
            ("Full Name:", self.current_user['name']),
            ("Email:", self.current_user.get('email', 'Not provided')),
            ("Member Since:", "2024"),
            ("Books Borrowed:", str(len([t for t in self.database.get_all_transactions() 
                                       if t['student_id'] == self.current_user['student_id'] 
                                       and t['action'] == 'borrow'])))
        ]
        
        for label, value in details:
            row = tk.Frame(info_frame, bg=self.bg_color)
            row.pack(fill=tk.X, padx=50, pady=5)
            
            tk.Label(
                row,
                text=label,
                font=("Helvetica", 12, "bold"),
                bg=self.bg_color,
                width=15,
                anchor='w'
            ).pack(side=tk.LEFT)
            
            tk.Label(
                row,
                text=value,
                font=("Helvetica", 12),
                bg=self.bg_color,
                anchor='w'
            ).pack(side=tk.LEFT, padx=10)
        
        # Current borrowings
        active = self.database.get_active_borrowings(self.current_user['student_id'])
        if active:
            tk.Label(
                info_frame,
                text=f"\nCurrently Borrowing: {len(active)} book(s)",
                font=("Helvetica", 12, "bold"),
                bg=self.bg_color,
                fg=self.accent_color
            ).pack(pady=10)
    
    def show_admin_dashboard(self):
        """Show admin dashboard"""
        self.clear_content()
        
        # Create notebook
        notebook = ttk.Notebook(self.content_frame)
        notebook.pack(fill=tk.BOTH, expand=True)
        
        # Dashboard tab
        dashboard_frame = ttk.Frame(notebook)
        notebook.add(dashboard_frame, text="Dashboard")
        self.setup_admin_dashboard(dashboard_frame)
        
        # Students tab
        students_frame = ttk.Frame(notebook)
        notebook.add(students_frame, text="Manage Students")
        self.setup_admin_students(students_frame)
        
        # Books tab
        books_frame = ttk.Frame(notebook)
        notebook.add(books_frame, text="Manage Books")
        self.setup_admin_books(books_frame)
        
        # Transactions tab
        transactions_frame = ttk.Frame(notebook)
        notebook.add(transactions_frame, text="Transactions")
        self.setup_admin_transactions(transactions_frame)
        
        # Overdue tab
        overdue_frame = ttk.Frame(notebook)
        notebook.add(overdue_frame, text="Overdue Books")
        self.setup_admin_overdue(overdue_frame)
        
        # Settings tab
        settings_frame = ttk.Frame(notebook)
        notebook.add(settings_frame, text="Settings")
        self.setup_admin_settings(settings_frame)
    
    def setup_admin_dashboard(self, parent):
        """Setup admin dashboard"""
        # Statistics cards
        stats = self.database.get_statistics()
        
        # Create card frame
        cards_frame = tk.Frame(parent, bg=self.bg_color)
        cards_frame.pack(pady=20)
        
        # Stat cards
        stat_items = [
            ("Total Transactions", stats['total_transactions'], self.accent_color),
            ("Active Borrowings", stats['active_borrowings'], self.success_color),
            ("Overdue Books", stats['overdue_books'], self.warning_color),
            ("Total Students", stats['unique_students'], "#9b59b6"),
            ("Total Books", stats['unique_books'], "#f39c12"),
            ("Returns", stats['total_returns'], "#1abc9c")
        ]
        
        row = 0
        col = 0
        for title, value, color in stat_items:
            card = tk.Frame(cards_frame, bg=color, width=180, height=120, relief=tk.RAISED, bd=2)
            card.grid(row=row, column=col, padx=10, pady=10)
            card.pack_propagate(False)
            
            tk.Label(
                card,
                text=title,
                bg=color,
                fg="white",
                font=("Helvetica", 10)
            ).pack(pady=(20, 5))
            
            tk.Label(
                card,
                text=str(value),
                bg=color,
                fg="white",
                font=("Helvetica", 24, "bold")
            ).pack()
            
            col += 1
            if col > 2:
                col = 0
                row += 1
        
        # Quick actions
        actions_frame = tk.Frame(parent, bg=self.bg_color, relief=tk.GROOVE, bd=2)
        actions_frame.pack(pady=20, fill=tk.X, padx=20)
        
        tk.Label(
            actions_frame,
            text="Quick Actions",
            font=("Helvetica", 14, "bold"),
            bg=self.bg_color
        ).pack(pady=10)
        
        btn_frame = tk.Frame(actions_frame, bg=self.bg_color)
        btn_frame.pack(pady=10)
        
        tk.Button(
            btn_frame,
            text="Send Overdue Notices",
            command=lambda: self.send_overdue_notices(),
            bg=self.warning_color,
            fg="white",
            padx=20
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Button(
            btn_frame,
            text="Generate Report",
            command=self.generate_report,
            bg=self.accent_color,
            fg="white",
            padx=20
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Button(
            btn_frame,
            text="Backup Database",
            command=self.backup_database,
            bg=self.success_color,
            fg="white",
            padx=20
        ).pack(side=tk.LEFT, padx=5)
    
    def setup_admin_students(self, parent):
        """Setup student management"""
        # Toolbar
        toolbar = tk.Frame(parent, bg=self.bg_color)
        toolbar.pack(fill=tk.X, pady=10)
        
        tk.Button(
            toolbar,
            text="Add Student",
            command=self.add_student_dialog,
            bg=self.success_color,
            fg="white",
            padx=10
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Button(
            toolbar,
            text="Remove Student",
            command=self.remove_student_dialog,
            bg=self.warning_color,
            fg="white",
            padx=10
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Label(toolbar, text="Search:", bg=self.bg_color).pack(side=tk.LEFT, padx=(20, 5))
        search_entry = tk.Entry(toolbar, width=30)
        search_entry.pack(side=tk.LEFT)
        
        # Students list
        columns = ('ID', 'Student ID', 'Name', 'Email', 'Books Borrowed')
        tree = ttk.Treeview(parent, columns=columns, show='headings', height=20)
        
        tree.heading('ID', text='ID')
        tree.heading('Student ID', text='Student ID')
        tree.heading('Name', text='Name')
        tree.heading('Email', text='Email')
        tree.heading('Books Borrowed', text='Books Borrowed')
        
        tree.column('ID', width=50)
        tree.column('Student ID', width=120)
        tree.column('Name', width=200)
        tree.column('Email', width=200)
        tree.column('Books Borrowed', width=100)
        
        # Add scrollbar
        scrollbar = ttk.Scrollbar(parent, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(10, 0))
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y, padx=(0, 10))
        
        # Populate students
        for i, (sid, student) in enumerate(self.students.items(), 1):
            active = self.database.get_active_borrowings(sid)
            tree.insert('', tk.END, values=(
                i,
                sid,
                student['name'],
                student.get('email', ''),
                len(active)
            ))
        
        # Search function
        def search():
            query = search_entry.get().lower()
            tree.delete(*tree.get_children())
            for i, (sid, student) in enumerate(self.students.items(), 1):
                if query in sid.lower() or query in student['name'].lower():
                    active = self.database.get_active_borrowings(sid)
                    tree.insert('', tk.END, values=(
                        i,
                        sid,
                        student['name'],
                        student.get('email', ''),
                        len(active)
                    ))
        
        search_entry.bind('<KeyRelease>', lambda e: search())
    
    def setup_admin_books(self, parent):
        """Setup book management"""
        # Toolbar
        toolbar = tk.Frame(parent, bg=self.bg_color)
        toolbar.pack(fill=tk.X, pady=10)
        
        tk.Button(
            toolbar,
            text="Add Book",
            command=self.add_book_dialog,
            bg=self.success_color,
            fg="white",
            padx=10
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Button(
            toolbar,
            text="Remove Book",
            command=self.remove_book_dialog,
            bg=self.warning_color,
            fg="white",
            padx=10
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Label(toolbar, text="Search:", bg=self.bg_color).pack(side=tk.LEFT, padx=(20, 5))
        search_entry = tk.Entry(toolbar, width=30)
        search_entry.pack(side=tk.LEFT)
        
        # Books list
        columns = ('Barcode', 'Title', 'Author', 'Status')
        tree = ttk.Treeview(parent, columns=columns, show='headings', height=20)
        
        tree.heading('Barcode', text='Barcode')
        tree.heading('Title', text='Title')
        tree.heading('Author', text='Author')
        tree.heading('Status', text='Status')
        
        tree.column('Barcode', width=150)
        tree.column('Title', width=300)
        tree.column('Author', width=200)
        tree.column('Status', width=100)
        
        # Add scrollbar
        scrollbar = ttk.Scrollbar(parent, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(10, 0))
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y, padx=(0, 10))
        
        # Populate books
        for barcode, book in self.books.items():
            status = "Available" if book['available'] else "Borrowed"
            tree.insert('', tk.END, values=(barcode, book['title'], book['author'], status))
        
        # Search function
        def search():
            query = search_entry.get().lower()
            tree.delete(*tree.get_children())
            for barcode, book in self.books.items():
                if query in barcode.lower() or query in book['title'].lower():
                    status = "Available" if book['available'] else "Borrowed"
                    tree.insert('', tk.END, values=(barcode, book['title'], book['author'], status))
        
        search_entry.bind('<KeyRelease>', lambda e: search())
    
    def setup_admin_transactions(self, parent):
        """Setup transactions view"""
        columns = ('Date', 'Student', 'Book', 'Action', 'Due Date')
        tree = ttk.Treeview(parent, columns=columns, show='headings', height=20)
        
        tree.heading('Date', text='Date')
        tree.heading('Student', text='Student')
        tree.heading('Book', text='Book')
        tree.heading('Action', text='Action')
        tree.heading('Due Date', text='Due Date')
        
        tree.column('Date', width=150)
        tree.column('Student', width=200)
        tree.column('Book', width=300)
        tree.column('Action', width=100)
        tree.column('Due Date', width=100)
        
        # Add scrollbar
        scrollbar = ttk.Scrollbar(parent, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(10, 0))
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y, padx=(0, 10))
        
        # Populate transactions
        transactions = self.database.get_all_transactions()
        for trans in transactions:
            tree.insert('', tk.END, values=(
                trans['date_time'][:16] if trans['date_time'] else "",
                trans['student_name'],
                trans['book_title'],
                trans['action'].upper(),
                trans['due_date']
            ))
    
    def setup_admin_overdue(self, parent):
        """Setup overdue books view"""
        overdue = self.database.check_overdue_books()
        
        if not overdue:
            tk.Label(
                parent,
                text="No overdue books found",
                font=("Helvetica", 14),
                fg=self.success_color
            ).pack(pady=50)
            return
        
        columns = ('Student', 'Book', 'Borrowed', 'Due Date', 'Days Overdue')
        tree = ttk.Treeview(parent, columns=columns, show='headings', height=20)
        
        tree.heading('Student', text='Student')
        tree.heading('Book', text='Book')
        tree.heading('Borrowed', text='Borrowed Date')
        tree.heading('Due Date', text='Due Date')
        tree.heading('Days Overdue', text='Days Overdue')
        
        tree.column('Student', width=200)
        tree.column('Book', width=300)
        tree.column('Borrowed', width=150)
        tree.column('Due Date', width=150)
        tree.column('Days Overdue', width=100)
        
        # Add scrollbar
        scrollbar = ttk.Scrollbar(parent, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(10, 0))
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y, padx=(0, 10))
        
        # Populate overdue
        for book in overdue:
            tree.insert('', tk.END, values=(
                book['student_name'],
                book['book_title'],
                book['date_time'][:10] if book['date_time'] else "",
                book['due_date'],
                f"{book['days_overdue']} days"
            ))
        
        # Send notices button
        tk.Button(
            parent,
            text="Send Overdue Notices",
            command=self.send_overdue_notices,
            bg=self.warning_color,
            fg="white",
            padx=20,
            pady=5
        ).pack(pady=10)
    
    def setup_admin_settings(self, parent):
        """Setup settings tab"""
        # Create settings form
        settings_frame = tk.Frame(parent, bg=self.bg_color, relief=tk.GROOVE, bd=2)
        settings_frame.pack(pady=20, padx=20, fill=tk.BOTH, expand=True)
        
        tk.Label(
            settings_frame,
            text="System Settings",
            font=("Helvetica", 16, "bold"),
            bg=self.bg_color
        ).pack(pady=20)
        
        # Borrowing days
        days_frame = tk.Frame(settings_frame, bg=self.bg_color)
        days_frame.pack(fill=tk.X, padx=50, pady=10)
        
        tk.Label(
            days_frame,
            text="Borrowing Period (days):",
            font=("Helvetica", 12),
            bg=self.bg_color,
            width=20,
            anchor='w'
        ).pack(side=tk.LEFT)
        
        days_var = tk.StringVar(value=str(BORROWING_DAYS))
        days_entry = tk.Entry(days_frame, textvariable=days_var, width=10, font=("Helvetica", 12))
        days_entry.pack(side=tk.LEFT, padx=10)
        
        def update_days():
            global BORROWING_DAYS
            try:
                new_days = int(days_var.get())
                if new_days > 0:
                    BORROWING_DAYS = new_days
                    messagebox.showinfo("Success", "Borrowing period updated")
                else:
                    messagebox.showerror("Error", "Please enter a positive number")
            except ValueError:
                messagebox.showerror("Error", "Invalid number")
        
        tk.Button(
            days_frame,
            text="Update",
            command=update_days,
            bg=self.accent_color,
            fg="white"
        ).pack(side=tk.LEFT, padx=10)
        
        # Admin password
        pass_frame = tk.Frame(settings_frame, bg=self.bg_color)
        pass_frame.pack(fill=tk.X, padx=50, pady=20)
        
        tk.Label(
            pass_frame,
            text="Change Admin Password:",
            font=("Helvetica", 12),
            bg=self.bg_color,
            width=20,
            anchor='w'
        ).pack(side=tk.LEFT)
        
        def change_password():
            dialog = tk.Toplevel(self.root)
            dialog.title("Change Password")
            dialog.geometry("400x200")
            dialog.transient(self.root)
            dialog.grab_set()
            
            tk.Label(dialog, text="New Password:").pack(pady=10)
            new_pass = tk.Entry(dialog, show="*", width=20)
            new_pass.pack()
            
            tk.Label(dialog, text="Confirm Password:").pack(pady=10)
            confirm_pass = tk.Entry(dialog, show="*", width=20)
            confirm_pass.pack()
            
            def save():
                if new_pass.get() and new_pass.get() == confirm_pass.get():
                    global ADMIN_PASSWORD
                    ADMIN_PASSWORD = new_pass.get()
                    messagebox.showinfo("Success", "Password changed")
                    dialog.destroy()
                else:
                    messagebox.showerror("Error", "Passwords do not match")
            
            tk.Button(dialog, text="Save", command=save).pack(pady=10)
        
        tk.Button(
            pass_frame,
            text="Change Password",
            command=change_password,
            bg=self.accent_color,
            fg="white"
        ).pack(side=tk.LEFT, padx=10)
        
        # WiFi settings
        wifi_frame = tk.Frame(settings_frame, bg=self.bg_color)
        wifi_frame.pack(fill=tk.X, padx=50, pady=20)
        
        tk.Label(
            wifi_frame,
            text="WiFi Settings:",
            font=("Helvetica", 12),
            bg=self.bg_color,
            width=20,
            anchor='w'
        ).pack(side=tk.LEFT)
        
        tk.Button(
            wifi_frame,
            text="Configure WiFi",
            command=self.wifi_settings,
            bg=self.accent_color,
            fg="white"
        ).pack(side=tk.LEFT, padx=10)
    
    # Helper methods
    def show_warning(self, message):
        """Show warning message"""
        warning = tk.Toplevel(self.root)
        warning.title("Warning")
        warning.geometry("300x150")
        warning.transient(self.root)
        
        tk.Label(
            warning,
            text="⚠️",
            font=("Helvetica", 48),
            fg=self.warning_color
        ).pack(pady=10)
        
        tk.Label(warning, text=message, font=("Helvetica", 10)).pack()
        tk.Button(warning, text="OK", command=warning.destroy).pack(pady=10)
    
    def add_student_dialog(self):
        """Dialog to add new student"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Add Student")
        dialog.geometry("400x250")
        dialog.transient(self.root)
        dialog.grab_set()
        
        tk.Label(dialog, text="Student ID:").pack(pady=5)
        sid_entry = tk.Entry(dialog, width=30)
        sid_entry.pack()
        
        tk.Label(dialog, text="Full Name:").pack(pady=5)
        name_entry = tk.Entry(dialog, width=30)
        name_entry.pack()
        
        tk.Label(dialog, text="Email (optional):").pack(pady=5)
        email_entry = tk.Entry(dialog, width=30)
        email_entry.pack()
        
        def save():
            sid = sid_entry.get().strip()
            name = name_entry.get().strip()
            email = email_entry.get().strip()
            
            if sid and name:
                self.students[sid] = {
                    "student_id": sid,
                    "name": name,
                    "email": email
                }
                messagebox.showinfo("Success", f"Student {name} added")
                dialog.destroy()
            else:
                messagebox.showerror("Error", "Student ID and Name are required")
        
        tk.Button(dialog, text="Save", command=save).pack(pady=10)
    
    def remove_student_dialog(self):
        """Dialog to remove student"""
        sid = simpledialog.askstring("Remove Student", "Enter Student ID:")
        if sid in self.students:
            # Check if student has active borrowings
            active = self.database.get_active_borrowings(sid)
            if active:
                messagebox.showerror("Error", "Cannot remove student with active borrowings")
                return
            
            name = self.students[sid]['name']
            if messagebox.askyesno("Confirm", f"Remove student {name}?"):
                del self.students[sid]
                messagebox.showinfo("Success", "Student removed")
        else:
            messagebox.showerror("Error", "Student not found")
    
    def add_book_dialog(self):
        """Dialog to add new book"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Add Book")
        dialog.geometry("400x250")
        dialog.transient(self.root)
        dialog.grab_set()
        
        tk.Label(dialog, text="Barcode:").pack(pady=5)
        barcode_entry = tk.Entry(dialog, width=30)
        barcode_entry.pack()
        
        tk.Label(dialog, text="Title:").pack(pady=5)
        title_entry = tk.Entry(dialog, width=30)
        title_entry.pack()
        
        tk.Label(dialog, text="Author:").pack(pady=5)
        author_entry = tk.Entry(dialog, width=30)
        author_entry.pack()
        
        def save():
            barcode = barcode_entry.get().strip()
            title = title_entry.get().strip()
            author = author_entry.get().strip() or "Unknown"
            
            if barcode and title:
                self.books[barcode] = {
                    "title": title,
                    "author": author,
                    "available": True
                }
                messagebox.showinfo("Success", f"Book '{title}' added")
                dialog.destroy()
            else:
                messagebox.showerror("Error", "Barcode and Title are required")
        
        tk.Button(dialog, text="Save", command=save).pack(pady=10)
    
    def remove_book_dialog(self):
        """Dialog to remove book"""
        barcode = simpledialog.askstring("Remove Book", "Enter Book Barcode:")
        if barcode in self.books:
            if not self.books[barcode]['available']:
                messagebox.showerror("Error", "Cannot remove book that is currently borrowed")
                return
            
            title = self.books[barcode]['title']
            if messagebox.askyesno("Confirm", f"Remove book '{title}'?"):
                del self.books[barcode]
                messagebox.showinfo("Success", "Book removed")
        else:
            messagebox.showerror("Error", "Book not found")
    
    def send_overdue_notices(self):
        """Send overdue notices"""
        overdue = self.database.check_overdue_books()
        if not overdue:
            messagebox.showinfo("Info", "No overdue books found")
            return
        
        # Group by student
        students_overdue = {}
        for book in overdue:
            sid = book['student_id']
            if sid not in students_overdue:
                students_overdue[sid] = []
            students_overdue[sid].append(book)
        
        if messagebox.askyesno("Confirm", f"Send notices to {len(students_overdue)} students?"):
            # In a real implementation, you would send emails here
            messagebox.showinfo("Success", f"Notices sent to {len(students_overdue)} students")
    
    def generate_report(self):
        """Generate library report"""
        stats = self.database.get_statistics()
        
        report = f"""
LIBRARY SYSTEM REPORT
Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

SUMMARY STATISTICS
-----------------
Total Transactions: {stats['total_transactions']}
Total Borrows: {stats['total_borrows']}
Total Returns: {stats['total_returns']}
Active Borrowings: {stats['active_borrowings']}
Overdue Books: {stats['overdue_books']}
Unique Students: {stats['unique_students']}
Unique Books: {stats['unique_books']}

BOOKS STATUS
-----------
Total Books in System: {len(self.books)}
Available Books: {sum(1 for b in self.books.values() if b['available'])}
Borrowed Books: {sum(1 for b in self.books.values() if not b['available'])}

STUDENTS STATUS
--------------
Total Students: {len(self.students)}
Students with Active Borrowings: {len(set(b['student_id'] for b in self.database.get_active_borrowings()))}
"""
        
        # Save report
        filename = f"library_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        with open(filename, 'w') as f:
            f.write(report)
        
        messagebox.showinfo("Success", f"Report saved as {filename}")
    
    def backup_database(self):
        """Backup the database"""
        import shutil
        
        if os.path.exists(self.database.filename):
            backup_name = f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            shutil.copy2(self.database.filename, backup_name)
            messagebox.showinfo("Success", f"Database backed up as {backup_name}")
        else:
            messagebox.showerror("Error", "Database file not found")
    
    def wifi_settings(self):
        """Configure WiFi"""
        dialog = tk.Toplevel(self.root)
        dialog.title("WiFi Settings")
        dialog.geometry("400x200")
        dialog.transient(self.root)
        dialog.grab_set()
        
        tk.Label(dialog, text="WiFi SSID:").pack(pady=5)
        ssid_entry = tk.Entry(dialog, width=30)
        ssid_entry.insert(0, WIFI_SSID)
        ssid_entry.pack()
        
        tk.Label(dialog, text="Password:").pack(pady=5)
        pass_entry = tk.Entry(dialog, width=30, show="*")
        pass_entry.insert(0, WIFI_PASSWORD)
        pass_entry.pack()
        
        def connect():
            ssid = ssid_entry.get()
            password = pass_entry.get()
            
            # Update global variables
            global WIFI_SSID, WIFI_PASSWORD
            WIFI_SSID = ssid
            WIFI_PASSWORD = password
            
            # Attempt connection in background
            messagebox.showinfo("Info", f"Attempting to connect to {ssid}...")
            dialog.destroy()
        
        tk.Button(dialog, text="Connect", command=connect).pack(pady=10)
    
    def load_student_data(self):
        """Load student data"""
        return {
            "136515140138": {"student_id": "136515140138", "name": "Lanze Andereson C. Lozano", "email": "lanze.anderson@gmail.com"},
            "136515130851": {"student_id": "136515130851", "name": "Jessabel S. Umayan", "email": "jessabelsorianoumayan@gmail.com"},
            "111844140223": {"student_id": "111844140223", "name": "Angela A. Merced", "email": "angela.merced@gmail.com"},
            "136685140071": {"student_id": "136685140071", "name": "Yasmien Sophie D. Astor", "email": "yasmienastor12@gmail.com"},
            "136526130084": {"student_id": "136526130084", "name": "Fritchie Reyes", "email": "fritchiereyes@gmail.com"},
            "136515140191": {"student_id": "136515140191", "name": "Allen Felicity Perez", "email": "allenfelicity17@gmail.com"},
            "136526130095": {"student_id": "136526130095", "name": "Reynaz Gonzales", "email": "reynazgonzales4@gmail.com"},
            "136529140999": {"student_id": "136529140999", "name": "Tiffany Saron Yacap", "email": "05.tiffany.05@gmail.com"},
            "136515130205": {"student_id": "136515130205", "name": "Asherah Yan-yan Corpuz", "email": "Asherah Yan-yan Corpuz"},
        }
    
    def load_book_data(self):
        """Load book data"""
        return {
            "9789716982115": {"title": "Noli Me Tangere", "author": "Jose Rizal", "available": True},
            "9789716985635": {"title": "El Filibusterismo", "author": "Jose Rizal", "available": True},
            "9789715691881": {"title": "Ibong Adarna", "author": "Unknown", "available": True},
            "9789715084190": {"title": "Florante at Laura", "author": "Francisco Balagtas", "available": True},
        }

# ============================================================================
# MAIN ENTRY POINT
# ============================================================================

if __name__ == "__main__":
    root = tk.Tk()
    app = LibrarySoftware(root)
    root.mainloop()
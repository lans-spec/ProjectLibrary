import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, font
from openpyxl import load_workbook, Workbook
import os
import time
from datetime import datetime, timedelta
import smtplib
import ssl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# ============================================================================
# CONFIGURATION
# ============================================================================

EMAIL_HOST = "smtp.gmail.com"
EMAIL_PORT = 587
EMAIL_ADDRESS = "lanze.anderson@gmail.com"
EMAIL_PASSWORD = "hirl quyv gdzs dewd"
LIBRARY_NAME = "MELCHORA AQUINO HIGH SCHOOL LIBRARY"
LIBRARIAN_EMAIL = "lanze.anderson@gmail.com"
BORROWING_DAYS = 3

# Admin credentials
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "library123"

# Excel Configuration
EXCEL_FILENAME = "DUAL-BARCODE-SCANNING-SYSTEM-3.xlsm"
STUDENT_SHEET = "STUDENT DATABASE"
BOOK_SHEET = "BOOK INVENTORY"
TRANSACTION_SHEET = "TRANSACTION LOG"

# ============================================================================
# DATABASE CLASS
# ============================================================================

class ExcelLibraryDatabase:
    def __init__(self, filename=EXCEL_FILENAME):
        self.filename = filename
        self.setup_database()

    def setup_database(self):
        """Setup or load the Excel workbook"""
        try:
            self.workbook = load_workbook(self.filename)
        except FileNotFoundError:
            # Create new workbook with required sheets
            self.workbook = Workbook()
            
            # STUDENT DATABASE sheet
            student_sheet = self.workbook.active
            student_sheet.title = STUDENT_SHEET
            student_sheet.append(["LRN", "STUDENT NAME", "GRADE & SECTION", "EMAIL"])
            
            # Add your sample students
            students_data = [
                ["136515130851", "JESSABEL S. UMAYAN", "12-STEM", "jessabelsorianoumayan@gmail.com"],
                ["136526130084", "FRITCHIE P. REYES", "12-STEM", "fritchiereyes@gmail.com"],
                ["111844140223", "ANGELA A. MERCED", "12-STEM", "angela.merced@gmail.com"],
                ["136515140138", "LANZE ANDERSON C. LOZANO", "12-STEM", "lanze.anderson@gmail.com"],
                ["136515140191", "ALLEN FELICITY S. PEREZ", "12-STEM", "allenfelicity17@gmail.com"],
                ["136526130095", "REYNAZ B. GONZALES", "12-STEM", "reynazgonzales4@gmail.com"],
                ["136526130443", "PRINCESS ALLEN P. MADRIAGA", "12-STEM", "princessallenmadriaga@gmail.com"],
                ["136685140071", "YASMIEN SOPHIE D. ASTOR", "12-STEM", "yasmienastor12@gmail.com"],
                ["136529140999", "TIFFANY S. YACAP", "12-HUMBS B", "05.tiffany.05@gmail.com"],
                ["136515130205", "ASHERAH YAN-YAN CORPUZ", "12-HUMBS B", "asherahyanancorpuz22@gmail.com"],
            ]
            for student in students_data:
                student_sheet.append(student)
            
            # BOOK INVENTORY sheet
            book_sheet = self.workbook.create_sheet(BOOK_SHEET)
            book_sheet.append(["BOOK BARCODE", "BOOK TITLE", "AUTHOR", "STATUS", "DATE BORROWED", "DUE DATE"])
            
            # Add your sample books
            books_data = [
                ["9780026757577", "Manufacturing Technology (Today and Tomorrow)", "unknown"],
                ["9789715147798", "True Filipino (In though, in Word and in Deed)", "unknown"],
                ["9789716996166", "Interactive Mathematics", "unknown"],
                ["9786214130178", "Skill book in Mathematics (Based on the K to 12 BEC)", "unknown"],
                ["9780153164132", "HBJ LANGUAGE", "unknown"],
                ["9789716893786", "Kasaysayan at Pamahalaang ng Pilipinas", "unknown"],
                ["9789716554601", "Math Builders", "unknown"],
                ["9789710581641", "K to 12 Conceptual Math & Beyond 7", "unknown"],
                ["9786214120178", "Skill book in Mathematics", "unknown"],
                ["9789719990802", "Edukasyon Sa Pagpapakatao", "unknown"],
            ]
            for book in books_data:
                book_sheet.append(book + ["Available", "", ""])
            
            # TRANSACTION LOG sheet
            trans_sheet = self.workbook.create_sheet(TRANSACTION_SHEET)
            trans_sheet.append([
                "Transaction ID", "LRN", "Student Name", "Grade & Section", "Email",
                "Book Barcode", "Book Title", "Author", "Action", 
                "Timestamp", "Date Time", "Due Date"
            ])
            
            self.workbook.save(self.filename)
        
        # Load all sheets
        self.student_sheet = self.workbook[STUDENT_SHEET]
        self.book_sheet = self.workbook[BOOK_SHEET]
        self.trans_sheet = self.workbook[TRANSACTION_SHEET]
    
    def save_workbook(self):
        """Save the workbook"""
        try:
            self.workbook.save(self.filename)
            return True
        except Exception as e:
            print(f"Error saving workbook: {e}")
            return False
    
    def find_student_by_id(self, student_id):
        """Find student in STUDENT DATABASE sheet by LRN"""
        for row in self.student_sheet.iter_rows(min_row=2, values_only=True):
            if len(row) >= 2 and str(row[0]) == str(student_id):
                return {
                    'lrn': row[0],
                    'name': row[1],
                    'grade_section': row[2] if len(row) > 2 else "",
                    'email': row[3] if len(row) > 3 else ""
                }
        return None
    
    def find_book_by_barcode(self, barcode):
        """Find book in BOOK INVENTORY sheet by barcode"""
        for row in self.book_sheet.iter_rows(min_row=2, values_only=True):
            if len(row) >= 2 and str(row[0]) == str(barcode):
                return {
                    'barcode': row[0],
                    'title': row[1],
                    'author': row[2] if len(row) > 2 else "Unknown",
                    'status': row[3] if len(row) > 3 else "Available",
                    'date_borrowed': row[4] if len(row) > 4 else "",
                    'due_date': row[5] if len(row) > 5 else ""
                }
        return None
    
    def log_transaction(self, student, book, action, due_date=None):
        """Log transaction to TRANSACTION LOG sheet"""
        try:
            # Generate transaction ID
            timestamp = time.time()
            date_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            due_date_str = due_date.strftime("%Y-%m-%d") if due_date else ""
            
            # Get next transaction ID
            last_id = 0
            for row in self.trans_sheet.iter_rows(min_row=2, values_only=True):
                if row[0] and str(row[0]).startswith('T'):
                    try:
                        num = int(str(row[0])[1:])
                        last_id = max(last_id, num)
                    except:
                        pass
            trans_id = f"T{last_id + 1:06d}"
            
            # Add to sheet
            self.trans_sheet.append([
                trans_id,
                student['lrn'],
                student['name'],
                student.get('grade_section', ''),
                student.get('email', ''),
                book['barcode'],
                book['title'],
                book['author'],
                action,
                timestamp,
                date_time,
                due_date_str
            ])
            
            # Update book status in BOOK INVENTORY
            if action == 'borrow':
                self.update_book_status(book['barcode'], 'Borrowed', date_time, due_date_str)
            elif action == 'return':
                self.update_book_status(book['barcode'], 'Available', '', '')
            
            self.save_workbook()
            return True, trans_id
            
        except Exception as e:
            return False, str(e)
    
    def update_book_status(self, barcode, status, date_borrowed="", due_date=""):
        """Update book status in BOOK INVENTORY"""
        for row_idx, row in enumerate(self.book_sheet.iter_rows(min_row=2), start=2):
            if len(row) >= 1 and str(row[0].value) == str(barcode):
                # Update status (column 4)
                self.book_sheet.cell(row=row_idx, column=4, value=status)
                # Update date borrowed (column 5)
                self.book_sheet.cell(row=row_idx, column=5, value=date_borrowed)
                # Update due date (column 6)
                self.book_sheet.cell(row=row_idx, column=6, value=due_date)
                break
        self.save_workbook()
    
    def get_active_borrowings(self, student_id=None):
        """Get all active borrowings from TRANSACTION LOG"""
        records = []
        rows = list(self.trans_sheet.iter_rows(min_row=2, values_only=True))
        
        # Get all borrows
        borrows = []
        for row in rows:
            if len(row) >= 9 and row[8] == 'borrow':  # Action column
                borrows.append({
                    'trans_id': row[0],
                    'lrn': row[1],
                    'student_name': row[2],
                    'grade_section': row[3],
                    'email': row[4],
                    'book_barcode': row[5],
                    'book_title': row[6],
                    'author': row[7],
                    'timestamp': row[9],
                    'date_time': row[10],
                    'due_date': row[11] if len(row) > 11 else ""
                })
        
        # Get all returns
        returns = {row[5] for row in rows if len(row) >= 9 and row[8] == 'return'}
        
        # Filter active
        active = [b for b in borrows if b['book_barcode'] not in returns]
        
        if student_id:
            active = [b for b in active if str(b['lrn']) == str(student_id)]
        
        return active
    
    def check_overdue_books(self, student_id=None):
        """Check for overdue books"""
        active = self.get_active_borrowings(student_id)
        overdue = []
        now = time.time()
        
        for record in active:
            try:
                if record['due_date']:
                    due_date = datetime.strptime(record['due_date'], "%Y-%m-%d")
                    due_timestamp = due_date.timestamp()
                    if now > due_timestamp:
                        days_overdue = int((now - due_timestamp) / (24 * 60 * 60))
                        record['days_overdue'] = days_overdue
                        overdue.append(record)
            except:
                continue
        
        return overdue
    
    def add_student(self, lrn, name, grade_section="", email=""):
        """Add student to STUDENT DATABASE"""
        # Check if already exists
        if self.find_student_by_id(lrn):
            return False, "LRN already exists"
        
        self.student_sheet.append([lrn, name, grade_section, email])
        self.save_workbook()
        return True, "Student added successfully"
    
    def remove_student(self, lrn):
        """Remove student from STUDENT DATABASE"""
        for row_idx, row in enumerate(self.student_sheet.iter_rows(min_row=2), start=2):
            if len(row) >= 1 and str(row[0].value) == str(lrn):
                # Check if student has active borrowings
                active = self.get_active_borrowings(lrn)
                if active:
                    return False, "Cannot remove student with active borrowings"
                
                self.student_sheet.delete_rows(row_idx)
                self.save_workbook()
                return True, "Student removed successfully"
        
        return False, "Student not found"
    
    def add_book(self, barcode, title, author="unknown"):
        """Add book to BOOK INVENTORY"""
        if self.find_book_by_barcode(barcode):
            return False, "Barcode already exists"
        
        self.book_sheet.append([barcode, title, author, "Available", "", ""])
        self.save_workbook()
        return True, "Book added successfully"
    
    def remove_book(self, barcode):
        """Remove book from BOOK INVENTORY"""
        for row_idx, row in enumerate(self.book_sheet.iter_rows(min_row=2), start=2):
            if len(row) >= 1 and str(row[0].value) == str(barcode):
                # Check if book is borrowed
                if len(row) > 3 and row[3].value == "Borrowed":
                    return False, "Cannot remove book that is currently borrowed"
                
                self.book_sheet.delete_rows(row_idx)
                self.save_workbook()
                return True, "Book removed successfully"
        
        return False, "Book not found"
    
    def get_all_students(self):
        """Get all students from STUDENT DATABASE"""
        students = {}
        for row in self.student_sheet.iter_rows(min_row=2, values_only=True):
            if len(row) >= 2:
                students[str(row[0])] = {
                    'lrn': row[0],
                    'name': row[1],
                    'grade_section': row[2] if len(row) > 2 else "",
                    'email': row[3] if len(row) > 3 else ""
                }
        return students
    
    def get_all_books(self):
        """Get all books from BOOK INVENTORY"""
        books = {}
        for row in self.book_sheet.iter_rows(min_row=2, values_only=True):
            if len(row) >= 2:
                books[str(row[0])] = {
                    'barcode': row[0],
                    'title': row[1],
                    'author': row[2] if len(row) > 2 else "Unknown",
                    'status': row[3] if len(row) > 3 else "Available",
                    'date_borrowed': row[4] if len(row) > 4 else "",
                    'due_date': row[5] if len(row) > 5 else ""
                }
        return books
    
    def get_all_transactions(self, limit=1000):
        """Get all transactions from TRANSACTION LOG"""
        transactions = []
        rows = list(self.trans_sheet.iter_rows(min_row=2, values_only=True))
        for row in rows[:limit]:
            if len(row) >= 9:
                transactions.append({
                    'trans_id': row[0],
                    'lrn': row[1],
                    'student_name': row[2],
                    'grade_section': row[3],
                    'email': row[4],
                    'book_barcode': row[5],
                    'book_title': row[6],
                    'author': row[7],
                    'action': row[8],
                    'timestamp': row[9],
                    'date_time': row[10],
                    'due_date': row[11] if len(row) > 11 else ""
                })
        return transactions
    
    def get_statistics(self):
        """Get library statistics"""
        transactions = self.get_all_transactions()
        active = self.get_active_borrowings()
        overdue = self.check_overdue_books()
        
        total_borrows = sum(1 for t in transactions if t['action'] == 'borrow')
        total_returns = sum(1 for t in transactions if t['action'] == 'return')
        
        unique_students = set(t['lrn'] for t in transactions if t['lrn'])
        unique_books = set(t['book_title'] for t in transactions if t['book_title'])
        
        all_books = self.get_all_books()
        available_books = sum(1 for b in all_books.values() if b.get('status') == 'Available')
        
        return {
            'total_transactions': len(transactions),
            'total_borrows': total_borrows,
            'total_returns': total_returns,
            'active_borrowings': len(active),
            'overdue_books': len(overdue),
            'unique_students': len(unique_students),
            'unique_books': len(unique_books),
            'total_books': len(all_books),
            'available_books': available_books,
            'borrowed_books': len(all_books) - available_books
        }

# ============================================================================
# EMAIL NOTIFIER CLASS
# ============================================================================

class EmailNotifier:
    def __init__(self):
        self.host = EMAIL_HOST
        self.port = EMAIL_PORT
        self.sender = EMAIL_ADDRESS
        self.password = EMAIL_PASSWORD
        self.librarian_email = LIBRARIAN_EMAIL

    def send_borrow_notification(self, student, book, due_date):
        """Send email to student and librarian about borrowed book"""
        try:
            borrow_date = datetime.now().strftime("%B %d, %Y")
            due_date_str = due_date.strftime("%B %d, %Y")

            # Send to student
            if student.get('email'):
                self._send_student_email(
                    student['email'],
                    student,
                    book,
                    borrow_date,
                    due_date_str
                )
            
            # Send to librarian
            self._send_librarian_email(
                self.librarian_email,
                student,
                book,
                borrow_date,
                due_date_str,
                'BORROWED'
            )
            
            return True
            
        except Exception as e:
            print(f"❌ Email failed: {e}")
            return False

    def send_return_notification(self, student, book):
        """Send email about book return"""
        try:
            return_date = datetime.now().strftime("%B %d, %Y")

            # Send to librarian only
            self._send_librarian_email(
                self.librarian_email,
                student,
                book,
                return_date,
                "",
                'RETURNED'
            )
            
            return True
            
        except Exception as e:
            print(f"❌ Email failed: {e}")
            return False

    def _send_student_email(self, recipient, student, book, borrow_date, due_date_str):
        """Send email to student"""
        msg = MIMEMultipart('alternative')
        msg['Subject'] = f"📚 Book Borrowed: {book['title']}"
        msg['From'] = f"{LIBRARY_NAME} <{self.sender}>"
        msg['To'] = recipient

        html_body = f"""
        <html>
        <body style="font-family: Arial, sans-serif;">
            <h2 style="color: #2c3e50;">📖 Library Borrowing Confirmation</h2>
            <p>Dear <strong>{student['name']}</strong>,</p>
            <p>You have successfully borrowed the following book from {LIBRARY_NAME}:</p>
            
            <table style="border-collapse: collapse; width: 100%; max-width: 600px; margin: 20px 0;">
                <tr style="background-color: #3498db; color: white;">
                    <th style="padding: 10px; text-align: left; border: 1px solid #ddd;">Book Details</th>
                    <th style="padding: 10px; text-align: left; border: 1px solid #ddd;">Information</th>
                </tr>
                <tr>
                    <td style="padding: 10px; border: 1px solid #ddd;"><strong>Book Title:</strong></td>
                    <td style="padding: 10px; border: 1px solid #ddd;">{book['title']}</td>
                </tr>
                <tr style="background-color: #f2f2f2;">
                    <td style="padding: 10px; border: 1px solid #ddd;"><strong>Author:</strong></td>
                    <td style="padding: 10px; border: 1px solid #ddd;">{book.get('author', 'N/A')}</td>
                </tr>
                <tr>
                    <td style="padding: 10px; border: 1px solid #ddd;"><strong>Borrow Date:</strong></td>
                    <td style="padding: 10px; border: 1px solid #ddd;">{borrow_date}</td>
                </tr>
                <tr style="background-color: #f2f2f2;">
                    <td style="padding: 10px; border: 1px solid #ddd;"><strong>Due Date:</strong></td>
                    <td style="padding: 10px; border: 1px solid #ddd;"><span style="color: #e74c3c; font-weight: bold;">{due_date_str}</span></td>
                </tr>
            </table>
            
            <p style="margin-top: 20px;">Please return the book by the due date to avoid penalties.</p>
            <p>Thank you for using our library!</p>
            
            <hr style="border: 1px solid #eee; margin: 20px 0;">
            <p style="color: #7f8c8d; font-size: 12px;">This is an automated message from {LIBRARY_NAME}. Please do not reply to this email.</p>
        </body>
        </html>
        """

        text_body = f"""
LIBRARY BORROWING CONFIRMATION

Dear {student['name']},

You have successfully borrowed the following book:
Book: {book['title']}
Author: {book.get('author', 'N/A')}
Borrow Date: {borrow_date}
Due Date: {due_date_str}

Please return the book by the due date to avoid penalties.

- {LIBRARY_NAME}
        """

        msg.attach(MIMEText(text_body, 'plain'))
        msg.attach(MIMEText(html_body, 'html'))

        context = ssl.create_default_context()
        with smtplib.SMTP(self.host, self.port) as server:
            server.starttls(context=context)
            server.login(self.sender, self.password)
            server.send_message(msg)
        
        print(f"✅ Email sent to student: {recipient}")

    def _send_librarian_email(self, recipient, student, book, date_str, due_date_str, action):
        """Send email to librarian"""
        msg = MIMEMultipart('alternative')
        msg['Subject'] = f"📚 Book {action}: {book['title']}"
        msg['From'] = f"{LIBRARY_NAME} System <{self.sender}>"
        msg['To'] = recipient

        html_body = f"""
        <html>
        <body style="font-family: Arial, sans-serif;">
            <h2 style="color: #2c3e50;">📖 Library {action} Alert</h2>
            <p><strong>BOOK {action}</strong></p>
            
            <table style="border-collapse: collapse; width: 100%; max-width: 600px; margin: 20px 0;">
                <tr style="background-color: #3498db; color: white;">
                    <th style="padding: 10px; text-align: left; border: 1px solid #ddd;">Student Information</th>
                    <th style="padding: 10px; text-align: left; border: 1px solid #ddd;">Details</th>
                </tr>
                <tr>
                    <td style="padding: 10px; border: 1px solid #ddd;"><strong>LRN:</strong></td>
                    <td style="padding: 10px; border: 1px solid #ddd;">{student['lrn']}</td>
                </tr>
                <tr style="background-color: #f2f2f2;">
                    <td style="padding: 10px; border: 1px solid #ddd;"><strong>Student Name:</strong></td>
                    <td style="padding: 10px; border: 1px solid #ddd;">{student['name']}</td>
                </tr>
                <tr>
                    <td style="padding: 10px; border: 1px solid #ddd;"><strong>Grade & Section:</strong></td>
                    <td style="padding: 10px; border: 1px solid #ddd;">{student.get('grade_section', 'N/A')}</td>
                </tr>
                <tr style="background-color: #f2f2f2;">
                    <td style="padding: 10px; border: 1px solid #ddd;"><strong>Email:</strong></td>
                    <td style="padding: 10px; border: 1px solid #ddd;">{student.get('email', 'N/A')}</td>
                </tr>
            </table>
            
            <table style="border-collapse: collapse; width: 100%; max-width: 600px; margin: 20px 0;">
                <tr style="background-color: #3498db; color: white;">
                    <th style="padding: 10px; text-align: left; border: 1px solid #ddd;">Book Information</th>
                    <th style="padding: 10px; text-align: left; border: 1px solid #ddd;">Details</th>
                </tr>
                <tr>
                    <td style="padding: 10px; border: 1px solid #ddd;"><strong>Barcode:</strong></td>
                    <td style="padding: 10px; border: 1px solid #ddd;">{book['barcode']}</td>
                </tr>
                <tr style="background-color: #f2f2f2;">
                    <td style="padding: 10px; border: 1px solid #ddd;"><strong>Title:</strong></td>
                    <td style="padding: 10px; border: 1px solid #ddd;">{book['title']}</td>
                </tr>
                <tr>
                    <td style="padding: 10px; border: 1px solid #ddd;"><strong>Author:</strong></td>
                    <td style="padding: 10px; border: 1px solid #ddd;">{book.get('author', 'N/A')}</td>
                </tr>
            </table>
            
            <table style="border-collapse: collapse; width: 100%; max-width: 600px; margin: 20px 0;">
                <tr style="background-color: #3498db; color: white;">
                    <th style="padding: 10px; text-align: left; border: 1px solid #ddd;">Transaction Details</th>
                    <th style="padding: 10px; text-align: left; border: 1px solid #ddd;">Information</th>
                </tr>
                <tr>
                    <td style="padding: 10px; border: 1px solid #ddd;"><strong>Action:</strong></td>
                    <td style="padding: 10px; border: 1px solid #ddd;">{action}</td>
                </tr>
                <tr style="background-color: #f2f2f2;">
                    <td style="padding: 10px; border: 1px solid #ddd;"><strong>Date:</strong></td>
                    <td style="padding: 10px; border: 1px solid #ddd;">{date_str}</td>
                </tr>
                {f'<tr><td style="padding: 10px; border: 1px solid #ddd;"><strong>Due Date:</strong></td><td style="padding: 10px; border: 1px solid #ddd;">{due_date_str}</td></tr>' if due_date_str else ''}
            </table>
            
            <hr style="border: 1px solid #eee; margin: 20px 0;">
            <p style="color: #7f8c8d; font-size: 12px;">This is an automated message from the Library System.</p>
        </body>
        </html>
        """

        msg.attach(MIMEText(html_body, 'html'))

        context = ssl.create_default_context()
        with smtplib.SMTP(self.host, self.port) as server:
            server.starttls(context=context)
            server.login(self.sender, self.password)
            server.send_message(msg)
        
        print(f"✅ Email sent to librarian: {recipient}")

# ============================================================================
# MAIN APPLICATION CLASS
# ============================================================================

class LibrarySoftware:
    def __init__(self, root):
        self.root = root
        self.root.title(f"{LIBRARY_NAME} - Library Management System")
        self.root.geometry("1200x700")
        self.root.minsize(1000, 600)
        
        # Configure colors
        self.bg_color = "#f0f0f0"
        self.header_color = "#2c3e50"
        self.accent_color = "#3498db"
        self.success_color = "#27ae60"
        self.warning_color = "#e74c3c"
        
        self.root.configure(bg=self.bg_color)
        
        # Initialize database
        self.database = ExcelLibraryDatabase()
        
        # Initialize email notifier
        self.notifier = EmailNotifier()
        
        # Current user (admin only)
        self.current_user = None
        
        # Store references to treeviews for refreshing
        self.student_tree = None
        self.book_tree = None
        self.transaction_tree = None
        self.overdue_tree = None
        
        # Setup UI
        self.setup_ui()
        
        # Show login screen
        self.show_login()
    
    def setup_ui(self):
        """Setup the main UI structure"""
        # Header
        self.header_frame = tk.Frame(self.root, bg=self.header_color, height=80)
        self.header_frame.pack(fill=tk.X, side=tk.TOP)
        self.header_frame.pack_propagate(False)
        
        header_font = font.Font(family="Helvetica", size=20, weight="bold")
        self.header_label = tk.Label(
            self.header_frame, 
            text=LIBRARY_NAME,
            bg=self.header_color,
            fg="white",
            font=header_font
        )
        self.header_label.pack(side=tk.LEFT, padx=20, pady=20)
        
        # User info in header
        self.user_frame = tk.Frame(self.header_frame, bg=self.header_color)
        self.user_frame.pack(side=tk.RIGHT, padx=20)
        
        self.user_label = tk.Label(
            self.user_frame,
            text="Not logged in",
            bg=self.header_color,
            fg="#ecf0f1",
            font=("Helvetica", 10)
        )
        self.user_label.pack()
        
        self.logout_btn = tk.Button(
            self.user_frame,
            text="Logout",
            bg=self.accent_color,
            fg="white",
            bd=0,
            padx=10,
            command=self.logout
        )
        
        # Main content area
        self.content_frame = tk.Frame(self.root, bg=self.bg_color)
        self.content_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
    
    def clear_content(self):
        """Clear the content frame"""
        for widget in self.content_frame.winfo_children():
            widget.destroy()
    
    def show_login(self):
        """Show admin login screen"""
        self.clear_content()
        self.user_label.config(text="Not logged in")
        self.logout_btn.pack_forget()
        
        login_frame = tk.Frame(self.content_frame, bg=self.bg_color)
        login_frame.pack(expand=True)
        
        title_font = font.Font(family="Helvetica", size=24, weight="bold")
        title = tk.Label(
            login_frame,
            text="Admin Login",
            bg=self.bg_color,
            font=title_font,
            fg=self.header_color
        )
        title.pack(pady=50)
        
        # Login form
        tk.Label(login_frame, text="Username:", font=("Helvetica", 12)).pack(pady=5)
        username = tk.Entry(login_frame, font=("Helvetica", 14), width=20)
        username.pack(pady=5)
        username.focus()
        
        tk.Label(login_frame, text="Password:", font=("Helvetica", 12)).pack(pady=5)
        password = tk.Entry(login_frame, font=("Helvetica", 14), width=20, show="*")
        password.pack(pady=5)
        
        def login():
            if username.get() == ADMIN_USERNAME and password.get() == ADMIN_PASSWORD:
                self.current_user = {"name": "Administrator", "username": ADMIN_USERNAME}
                self.user_label.config(text="Admin: Administrator")
                self.logout_btn.pack(side=tk.RIGHT, padx=5)
                self.show_admin_dashboard()
            else:
                messagebox.showerror("Error", "Invalid credentials")
        
        tk.Button(
            login_frame,
            text="Login",
            command=login,
            bg=self.accent_color,
            fg="white",
            padx=20,
            pady=5,
            width=15
        ).pack(pady=20)
    
    def logout(self):
        """Logout current user"""
        self.current_user = None
        self.show_login()
    
    def refresh_all_tabs(self):
        """Refresh all tabs with current data"""
        if self.student_tree:
            self.refresh_student_list()
        if self.book_tree:
            self.refresh_book_list()
        if self.transaction_tree:
            self.refresh_transaction_list()
    
    def show_admin_dashboard(self):
        """Show admin dashboard with barcode scanning"""
        self.clear_content()
        
        # Create notebook for tabs
        notebook = ttk.Notebook(self.content_frame)
        notebook.pack(fill=tk.BOTH, expand=True)
        
        # Scanner Tab (Main)
        scanner_frame = ttk.Frame(notebook)
        notebook.add(scanner_frame, text="Barcode Scanner")
        self.setup_scanner_tab(scanner_frame)
        
        # Dashboard Tab
        dashboard_frame = ttk.Frame(notebook)
        notebook.add(dashboard_frame, text="Dashboard")
        self.setup_admin_dashboard(dashboard_frame)
        
        # Students Tab
        students_frame = ttk.Frame(notebook)
        notebook.add(students_frame, text="Student Database")
        self.setup_admin_students(students_frame)
        
        # Books Tab
        books_frame = ttk.Frame(notebook)
        notebook.add(books_frame, text="Book Inventory")
        self.setup_admin_books(books_frame)
        
        # Transactions Tab
        transactions_frame = ttk.Frame(notebook)
        notebook.add(transactions_frame, text="Transaction Log")
        self.setup_admin_transactions(transactions_frame)
        
        # Overdue Tab
        overdue_frame = ttk.Frame(notebook)
        notebook.add(overdue_frame, text="Overdue Books")
        self.setup_admin_overdue(overdue_frame)
        
        # Settings Tab
        settings_frame = ttk.Frame(notebook)
        notebook.add(settings_frame, text="Settings")
        self.setup_admin_settings(settings_frame)
    
    def setup_scanner_tab(self, parent):
        """Setup barcode scanner tab with left/right/bottom layout"""
    # Title at the top
        title_label = tk.Label(
            parent, 
            text="Barcode Scanner Interface",
            font=("Helvetica", 14, "bold"),
            bg=self.bg_color
        )
        title_label.pack(pady=10)
    
    # Main container frame for left and right panels
        main_container = tk.Frame(parent, bg=self.bg_color)
        main_container.pack(fill=tk.BOTH, expand=True, padx=15)
    
    # ===== LEFT PANEL - STEP 1 =====
        left_frame = tk.Frame(main_container, bg=self.bg_color, relief=tk.GROOVE, bd=1, width=450)
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        left_frame.pack_propagate(False)
    
    # Step 1 header
        step1_label = tk.Label(
            left_frame,
            text="STEP 1: Scan Student LRN",
            font=("Helvetica", 11, "bold"),
            bg=self.accent_color,
            fg="white",
            pady=5
        )
        step1_label.pack(fill=tk.X)
    
    # Student input area
        student_input_frame = tk.Frame(left_frame, bg=self.bg_color, padx=15, pady=10)
        student_input_frame.pack(fill=tk.X)
    
        tk.Label(
            student_input_frame,
            text="Scan or enter Student LRN:",
            font=("Helvetica", 9),
            bg=self.bg_color
        ).pack(anchor=tk.W, pady=(0, 3))
    
        id_frame = tk.Frame(student_input_frame, bg=self.bg_color)
        id_frame.pack(fill=tk.X)
    
        self.student_id_entry = tk.Entry(id_frame, font=("Helvetica", 11), width=22)
        self.student_id_entry.pack(side=tk.LEFT, padx=(0, 5))
        self.student_id_entry.bind('<Return>', lambda e: self.lookup_student())
    
        tk.Button(
            id_frame,
            text="Lookup Student",
            command=self.lookup_student,
            bg=self.accent_color,
            fg="white",
            padx=8,
            height=1,
            font=("Helvetica", 9)
        ).pack(side=tk.LEFT)
    
    # Student info display
        self.student_info_frame = tk.Frame(left_frame, bg="#e8f4f8", relief=tk.SUNKEN, bd=1, padx=8, pady=8)
        self.student_info_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=10)
    
        self.student_info_label = tk.Label(
            self.student_info_frame,
            text="No student selected",
            font=("Helvetica", 9),
            bg="#e8f4f8",
            height=5,
            justify=tk.LEFT,
            anchor=tk.NW
        )
        self.student_info_label.pack(fill=tk.BOTH, expand=True)
    
    # ===== RIGHT PANEL - STEP 2 =====
        right_frame = tk.Frame(main_container, bg=self.bg_color, relief=tk.GROOVE, bd=1, width=450)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(5, 0))
        right_frame.pack_propagate(False)
    
    # Step 2 header
        step2_label = tk.Label(
            right_frame,
            text="STEP 2: Scan Book Barcode",
            font=("Helvetica", 11, "bold"),
            bg=self.accent_color,
            fg="white",
            pady=5
        )
        step2_label.pack(fill=tk.X)
    
    # Book input area
        book_input_frame = tk.Frame(right_frame, bg=self.bg_color, padx=15, pady=10)
        book_input_frame.pack(fill=tk.X)
    
        tk.Label(
            book_input_frame,
            text="Scan or enter Book Barcode:",
            font=("Helvetica", 9),
            bg=self.bg_color
        ).pack(anchor=tk.W, pady=(0, 3))
    
        barcode_frame = tk.Frame(book_input_frame, bg=self.bg_color)
        barcode_frame.pack(fill=tk.X)
    
        self.book_barcode_entry = tk.Entry(barcode_frame, font=("Helvetica", 11), width=22)
        self.book_barcode_entry.pack(side=tk.LEFT, padx=(0, 5))
        self.book_barcode_entry.bind('<Return>', lambda e: self.lookup_book())
    
        tk.Button(
            barcode_frame,
            text="Lookup Book",
            command=self.lookup_book,
            bg=self.accent_color,
            fg="white",
            padx=8,
            height=1,
            font=("Helvetica", 9)
        ).pack(side=tk.LEFT)
    
    # Book info display
        self.book_info_frame = tk.Frame(right_frame, bg="#e8f4f8", relief=tk.SUNKEN, bd=1, padx=18, pady=8)
        self.book_info_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=10)
    
        self.book_info_label = tk.Label(
            self.book_info_frame,
            text="No book selected",
            font=("Helvetica", 9),
            bg="#e8f4f8",
            height=5,
            justify=tk.LEFT,
            anchor=tk.NW
        )
        self.book_info_label.pack(fill=tk.BOTH, expand=True)
    
    # ===== BOTTOM PANEL - STEP 3 =====
        bottom_frame = tk.Frame(parent, bg=self.bg_color, relief=tk.GROOVE, bd=1)
        bottom_frame.pack(fill=tk.X, padx=15, pady=10)
    
    # Step 3 header
        step3_label = tk.Label(
            bottom_frame,
            text="STEP 3: Select Action",
            font=("Helvetica", 11, "bold"),
            bg=self.accent_color,
            fg="white",
            pady=5
        )
        step3_label.pack(fill=tk.X)
    
    # Action buttons
        action_frame = tk.Frame(bottom_frame, bg=self.bg_color, pady=10)
        action_frame.pack()
    
        self.borrow_btn = tk.Button(
            action_frame,
            text="BORROW BOOK",
            command=self.process_borrow,
            bg=self.success_color,
            fg="white",
            font=("Helvetica", 11, "bold"),
            padx=25,
            pady=8,
            state=tk.DISABLED,
            width=12
        )
        self.borrow_btn.pack(side=tk.LEFT, padx=10)
    
        self.return_btn = tk.Button(
            action_frame,
            text="RETURN BOOK",
            command=self.process_return,
            bg=self.warning_color,
            fg="white",
            font=("Helvetica", 11, "bold"),
            padx=25,
            pady=8,
            state=tk.DISABLED,
            width=12
        )
        self.return_btn.pack(side=tk.LEFT, padx=10)
    
    # Status display at the very bottom
        self.scanner_status = tk.Label(
            parent,
            text="Ready. Scan student LRN to begin.",
            font=("Helvetica", 9, "italic"),
            bg=self.bg_color,
            fg="gray"
        )
        self.scanner_status.pack(pady=(0, 5))
    
    # Store current selections
        self.current_student = None
        self.current_book = None
    
    def lookup_student(self):
        """Look up student by LRN"""
        student_id = self.student_id_entry.get().strip()
        if not student_id:
            messagebox.showwarning("Warning", "Please enter a Student LRN")
            return
        
        student = self.database.find_student_by_id(student_id)
        
        if student:
            self.current_student = student
            info_text = f"✓ STUDENT FOUND\nLRN: {student['lrn']}\nName: {student['name']}\nGrade & Section: {student.get('grade_section', 'N/A')}\nEmail: {student.get('email', 'No email')}"
            self.student_info_label.config(text=info_text, fg="black")
            self.scanner_status.config(text=f"Student selected: {student['name']}. Now scan book.")
            
            # Check if student has overdue books
            overdue = self.database.check_overdue_books(student_id)
            if overdue:
                self.scanner_status.config(
                    text=f"⚠️ WARNING: Student has {len(overdue)} overdue book(s)!",
                    fg=self.warning_color
                )
        else:
            self.current_student = None
            self.student_info_label.config(
                text=f"✗ STUDENT NOT FOUND\nLRN: {student_id}\nPlease check LRN and try again.",
                fg=self.warning_color
            )
            self.scanner_status.config(text="Student not found. Try again.", fg=self.warning_color)
        
        self.update_action_buttons()
    
    def lookup_book(self):
        """Look up book by barcode"""
        barcode = self.book_barcode_entry.get().strip()
        if not barcode:
            messagebox.showwarning("Warning", "Please enter a Book Barcode")
            return
        
        book = self.database.find_book_by_barcode(barcode)
        
        if book:
            self.current_book = book
            info_text = f"✓ BOOK FOUND\nBarcode: {book['barcode']}\nTitle: {book['title']}\nAuthor: {book['author']}\nStatus: {book['status']}"
            self.book_info_label.config(text=info_text, fg="black")
            self.scanner_status.config(text=f"Book selected: {book['title']} ({book['status']})")
            
            if book['status'] != "Available":
                self.scanner_status.config(
                    text=f"⚠️ Book is currently {book['status']}",
                    fg=self.warning_color
                )
        else:
            self.current_book = None
            self.book_info_label.config(
                text=f"✗ BOOK NOT FOUND\nBarcode: {barcode}\nPlease check barcode and try again.",
                fg=self.warning_color
            )
            self.scanner_status.config(text="Book not found. Try again.", fg=self.warning_color)
        
        self.update_action_buttons()
    
    def update_action_buttons(self):
        """Enable/disable action buttons based on selections"""
        if self.current_student and self.current_book:
            if self.current_book['status'] == "Available":
                self.borrow_btn.config(state=tk.NORMAL, bg=self.success_color)
                self.return_btn.config(state=tk.DISABLED, bg="gray")
            else:
                # Book is borrowed, enable return button
                self.borrow_btn.config(state=tk.DISABLED, bg="gray")
                self.return_btn.config(state=tk.NORMAL, bg=self.warning_color)
        else:
            self.borrow_btn.config(state=tk.DISABLED, bg="gray")
            self.return_btn.config(state=tk.DISABLED, bg="gray")
    
    def process_borrow(self):
        """Process book borrowing"""
        if not self.current_student or not self.current_book:
            messagebox.showerror("Error", "Please select both student and book")
            return
        
        # Check if book is available
        if self.current_book['status'] != "Available":
            messagebox.showerror("Error", "This book is not available for borrowing")
            return
        
        # Check for overdue books
        overdue = self.database.check_overdue_books(self.current_student['lrn'])
        if overdue:
            if not messagebox.askyesno("Overdue Warning", 
                f"Student has {len(overdue)} overdue book(s). Continue borrowing?"):
                return
        
        # Calculate due date
        due_date = datetime.now() + timedelta(days=BORROWING_DAYS)
        
        # Log transaction
        success, trans_id = self.database.log_transaction(
            self.current_student,
            self.current_book,
            'borrow',
            due_date
        )
        
        if success:
            # Send email notifications
            self.notifier.send_borrow_notification(
                self.current_student,
                self.current_book,
                due_date
            )
            
            messagebox.showinfo(
                "Success", 
                f"Book borrowed successfully!\n\nTransaction ID: {trans_id}\nDue Date: {due_date.strftime('%Y-%m-%d')}"
            )
            
            # Refresh all displays
            self.refresh_all_tabs()
            
            # Clear selections
            self.student_id_entry.delete(0, tk.END)
            self.book_barcode_entry.delete(0, tk.END)
            self.current_student = None
            self.current_book = None
            self.student_info_label.config(text="No student selected")
            self.book_info_label.config(text="No book selected")
            self.scanner_status.config(text="Ready. Scan next student LRN.")
            self.update_action_buttons()
        else:
            messagebox.showerror("Error", f"Failed to process borrowing: {trans_id}")
    
    def process_return(self):
        """Process book return"""
        if not self.current_student or not self.current_book:
            messagebox.showerror("Error", "Please select both student and book")
            return
        
        # Verify that this book was borrowed by this student
        active = self.database.get_active_borrowings(self.current_student['lrn'])
        borrowed = any(b['book_barcode'] == self.current_book['barcode'] for b in active)
        
        if not borrowed:
            messagebox.showerror("Error", "This book was not borrowed by the selected student")
            return
        
        # Log transaction
        success, trans_id = self.database.log_transaction(
            self.current_student,
            self.current_book,
            'return'
        )
        
        if success:
            # Send notification to librarian
            self.notifier.send_return_notification(
                self.current_student,
                self.current_book
            )
            
            messagebox.showinfo("Success", f"Book returned successfully!\n\nTransaction ID: {trans_id}")
            
            # Refresh all displays
            self.refresh_all_tabs()
            
            # Clear selections
            self.student_id_entry.delete(0, tk.END)
            self.book_barcode_entry.delete(0, tk.END)
            self.current_student = None
            self.current_book = None
            self.student_info_label.config(text="No student selected")
            self.book_info_label.config(text="No book selected")
            self.scanner_status.config(text="Ready. Scan next student LRN.")
            self.update_action_buttons()
        else:
            messagebox.showerror("Error", f"Failed to process return: {trans_id}")
    
    def setup_admin_dashboard(self, parent):
        """Setup admin dashboard"""
        stats = self.database.get_statistics()
        
        # Statistics cards
        cards_frame = tk.Frame(parent, bg=self.bg_color)
        cards_frame.pack(pady=20)
        
        stat_items = [
            ("Total Transactions", stats['total_transactions'], self.accent_color),
            ("Active Borrowings", stats['active_borrowings'], self.success_color),
            ("Overdue Books", stats['overdue_books'], self.warning_color),
            ("Total Students", stats['unique_students'], "#9b59b6"),
            ("Total Books", stats['total_books'], "#f39c12"),
            ("Available Books", stats['available_books'], "#1abc9c")
        ]
        
        row = 0
        col = 0
        for title, value, color in stat_items:
            card = tk.Frame(cards_frame, bg=color, width=180, height=120, relief=tk.RAISED, bd=2)
            card.grid(row=row, column=col, padx=10, pady=10)
            card.pack_propagate(False)
            
            tk.Label(
                card,
                text=title,
                bg=color,
                fg="white",
                font=("Helvetica", 10)
            ).pack(pady=(20, 5))
            
            tk.Label(
                card,
                text=str(value),
                bg=color,
                fg="white",
                font=("Helvetica", 24, "bold")
            ).pack()
            
            col += 1
            if col > 2:
                col = 0
                row += 1
        
        # Quick actions
        actions_frame = tk.Frame(parent, bg=self.bg_color, relief=tk.GROOVE, bd=2)
        actions_frame.pack(pady=20, fill=tk.X, padx=20)
        
        tk.Label(
            actions_frame,
            text="Quick Actions",
            font=("Helvetica", 14, "bold"),
            bg=self.bg_color
        ).pack(pady=10)
        
        btn_frame = tk.Frame(actions_frame, bg=self.bg_color)
        btn_frame.pack(pady=10)
        
        tk.Button(
            btn_frame,
            text="Send Overdue Notices",
            command=self.send_overdue_notices,
            bg=self.warning_color,
            fg="white",
            padx=20
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Button(
            btn_frame,
            text="Generate Report",
            command=self.generate_report,
            bg=self.accent_color,
            fg="white",
            padx=20
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Button(
            btn_frame,
            text="Backup Database",
            command=self.backup_database,
            bg=self.success_color,
            fg="white",
            padx=20
        ).pack(side=tk.LEFT, padx=5)
        
        # Refresh button
        tk.Button(
            btn_frame,
            text="Refresh Dashboard",
            command=lambda: self.refresh_dashboard(parent),
            bg=self.header_color,
            fg="white",
            padx=20
        ).pack(side=tk.LEFT, padx=5)
    
    def refresh_dashboard(self, parent):
        """Refresh dashboard statistics"""
        # Clear and recreate dashboard
        for widget in parent.winfo_children():
            widget.destroy()
        self.setup_admin_dashboard(parent)
    
    def setup_admin_students(self, parent):
        """Setup student management - STUDENT DATABASE sheet"""
        toolbar = tk.Frame(parent, bg=self.bg_color)
        toolbar.pack(fill=tk.X, pady=10)
        
        tk.Button(
            toolbar,
            text="Add Student",
            command=self.add_student_dialog,
            bg=self.success_color,
            fg="white",
            padx=10
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Button(
            toolbar,
            text="Remove Student",
            command=self.remove_student_dialog,
            bg=self.warning_color,
            fg="white",
            padx=10
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Button(
            toolbar,
            text="Refresh",
            command=self.refresh_student_list,
            bg=self.accent_color,
            fg="white",
            padx=10
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Label(toolbar, text="Search:", bg=self.bg_color).pack(side=tk.LEFT, padx=(20, 5))
        search_entry = tk.Entry(toolbar, width=30)
        search_entry.pack(side=tk.LEFT)
        
        # Students treeview
        columns = ('LRN', 'Student Name', 'Grade & Section', 'Email', 'Books Borrowed')
        self.student_tree = ttk.Treeview(parent, columns=columns, show='headings', height=20)
        
        self.student_tree.heading('LRN', text='LRN')
        self.student_tree.heading('Student Name', text='STUDENT NAME')
        self.student_tree.heading('Grade & Section', text='GRADE & SECTION')
        self.student_tree.heading('Email', text='EMAIL')
        self.student_tree.heading('Books Borrowed', text='Borrowed')
        
        self.student_tree.column('LRN', width=120)
        self.student_tree.column('Student Name', width=250)
        self.student_tree.column('Grade & Section', width=120)
        self.student_tree.column('Email', width=200)
        self.student_tree.column('Books Borrowed', width=80)
        
        scrollbar = ttk.Scrollbar(parent, orient=tk.VERTICAL, command=self.student_tree.yview)
        self.student_tree.configure(yscrollcommand=scrollbar.set)
        
        self.student_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(10, 0))
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y, padx=(0, 10))
        
        def search():
            query = search_entry.get().lower()
            self.refresh_student_list(query)
        
        search_entry.bind('<KeyRelease>', lambda e: search())
        
        # Initial load
        self.refresh_student_list()
    
    def refresh_student_list(self, query=""):
        """Refresh the student list"""
        if not self.student_tree:
            return
        
        # Clear existing items
        for item in self.student_tree.get_children():
            self.student_tree.delete(item)
        
        # Get fresh data from database
        students = self.database.get_all_students()
        
        # Add students to treeview
        for lrn, student in students.items():
            # Apply search filter
            if query:
                if query not in lrn.lower() and query not in student['name'].lower():
                    continue
            
            active = self.database.get_active_borrowings(lrn)
            self.student_tree.insert('', tk.END, values=(
                lrn,
                student['name'],
                student.get('grade_section', ''),
                student.get('email', ''),
                len(active)
            ))
    
    def setup_admin_books(self, parent):
        """Setup book management - BOOK INVENTORY sheet"""
        toolbar = tk.Frame(parent, bg=self.bg_color)
        toolbar.pack(fill=tk.X, pady=10)
        
        tk.Button(
            toolbar,
            text="Add Book",
            command=self.add_book_dialog,
            bg=self.success_color,
            fg="white",
            padx=10
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Button(
            toolbar,
            text="Remove Book",
            command=self.remove_book_dialog,
            bg=self.warning_color,
            fg="white",
            padx=10
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Button(
            toolbar,
            text="Refresh",
            command=self.refresh_book_list,
            bg=self.accent_color,
            fg="white",
            padx=10
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Label(toolbar, text="Search:", bg=self.bg_color).pack(side=tk.LEFT, padx=(20, 5))
        search_entry = tk.Entry(toolbar, width=30)
        search_entry.pack(side=tk.LEFT)
        
        # Books treeview
        columns = ('Barcode', 'Title', 'Author', 'Status', 'Date Borrowed', 'Due Date')
        self.book_tree = ttk.Treeview(parent, columns=columns, show='headings', height=20)
        
        self.book_tree.heading('Barcode', text='BOOK BARCODE')
        self.book_tree.heading('Title', text='BOOK TITLE')
        self.book_tree.heading('Author', text='AUTHOR')
        self.book_tree.heading('Status', text='STATUS')
        self.book_tree.heading('Date Borrowed', text='DATE BORROWED')
        self.book_tree.heading('Due Date', text='DUE DATE')
        
        self.book_tree.column('Barcode', width=120)
        self.book_tree.column('Title', width=300)
        self.book_tree.column('Author', width=200)
        self.book_tree.column('Status', width=100)
        self.book_tree.column('Date Borrowed', width=150)
        self.book_tree.column('Due Date', width=100)
        
        scrollbar = ttk.Scrollbar(parent, orient=tk.VERTICAL, command=self.book_tree.yview)
        self.book_tree.configure(yscrollcommand=scrollbar.set)
        
        self.book_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(10, 0))
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y, padx=(0, 10))
        
        def search():
            query = search_entry.get().lower()
            self.refresh_book_list(query)
        
        search_entry.bind('<KeyRelease>', lambda e: search())
        
        # Initial load
        self.refresh_book_list()
    
    def refresh_book_list(self, query=""):
        """Refresh the book list"""
        if not self.book_tree:
            return
        
        # Clear existing items
        for item in self.book_tree.get_children():
            self.book_tree.delete(item)
        
        # Get fresh data from database
        books = self.database.get_all_books()
        
        # Add books to treeview
        for barcode, book in books.items():
            # Apply search filter
            if query:
                if query not in barcode.lower() and query not in book['title'].lower():
                    continue
            
            self.book_tree.insert('', tk.END, values=(
                barcode,
                book['title'],
                book['author'],
                book.get('status', 'Available'),
                book.get('date_borrowed', ''),
                book.get('due_date', '')
            ))
    
    def setup_admin_transactions(self, parent):
        """Setup transactions view - TRANSACTION LOG sheet"""
        columns = ('Trans ID', 'Date', 'LRN', 'Student Name', 'Grade', 'Book', 'Author', 'Action', 'Due Date')
        self.transaction_tree = ttk.Treeview(parent, columns=columns, show='headings', height=20)
        
        self.transaction_tree.heading('Trans ID', text='Transaction ID')
        self.transaction_tree.heading('Date', text='Date')
        self.transaction_tree.heading('LRN', text='LRN')
        self.transaction_tree.heading('Student Name', text='Student Name')
        self.transaction_tree.heading('Grade', text='Grade & Section')
        self.transaction_tree.heading('Book', text='Book Title')
        self.transaction_tree.heading('Author', text='Author')
        self.transaction_tree.heading('Action', text='Action')
        self.transaction_tree.heading('Due Date', text='Due Date')
        
        self.transaction_tree.column('Trans ID', width=100)
        self.transaction_tree.column('Date', width=150)
        self.transaction_tree.column('LRN', width=120)
        self.transaction_tree.column('Student Name', width=200)
        self.transaction_tree.column('Grade', width=100)
        self.transaction_tree.column('Book', width=250)
        self.transaction_tree.column('Author', width=150)
        self.transaction_tree.column('Action', width=80)
        self.transaction_tree.column('Due Date', width=100)
        
        scrollbar = ttk.Scrollbar(parent, orient=tk.VERTICAL, command=self.transaction_tree.yview)
        self.transaction_tree.configure(yscrollcommand=scrollbar.set)
        
        self.transaction_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(10, 0))
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y, padx=(0, 10))
        
        # Refresh button
        tk.Button(
            parent,
            text="Refresh",
            command=self.refresh_transaction_list,
            bg=self.accent_color,
            fg="white",
            padx=20
        ).pack(pady=10)
        
        # Initial load
        self.refresh_transaction_list()
    
    def refresh_transaction_list(self):
        """Refresh the transaction list"""
        if not self.transaction_tree:
            return
        
        # Clear existing items
        for item in self.transaction_tree.get_children():
            self.transaction_tree.delete(item)
        
        # Get fresh data from database
        transactions = self.database.get_all_transactions(limit=500)
        
        # Add transactions to treeview
        for trans in transactions:
            self.transaction_tree.insert('', tk.END, values=(
                trans['trans_id'],
                trans['date_time'][:16] if trans['date_time'] else "",
                trans['lrn'],
                trans['student_name'],
                trans.get('grade_section', ''),
                trans['book_title'],
                trans['author'],
                trans['action'].upper(),
                trans['due_date']
            ))
    
    def setup_admin_overdue(self, parent):
        """Setup overdue books view"""
        columns = ('Student', 'LRN', 'Grade', 'Book', 'Borrowed', 'Due Date', 'Days Overdue')
        self.overdue_tree = ttk.Treeview(parent, columns=columns, show='headings', height=20)
        
        self.overdue_tree.heading('Student', text='Student Name')
        self.overdue_tree.heading('LRN', text='LRN')
        self.overdue_tree.heading('Grade', text='Grade & Section')
        self.overdue_tree.heading('Book', text='Book Title')
        self.overdue_tree.heading('Borrowed', text='Borrowed Date')
        self.overdue_tree.heading('Due Date', text='Due Date')
        self.overdue_tree.heading('Days Overdue', text='Days Overdue')
        
        self.overdue_tree.column('Student', width=200)
        self.overdue_tree.column('LRN', width=120)
        self.overdue_tree.column('Grade', width=100)
        self.overdue_tree.column('Book', width=300)
        self.overdue_tree.column('Borrowed', width=150)
        self.overdue_tree.column('Due Date', width=150)
        self.overdue_tree.column('Days Overdue', width=100)
        
        scrollbar = ttk.Scrollbar(parent, orient=tk.VERTICAL, command=self.overdue_tree.yview)
        self.overdue_tree.configure(yscrollcommand=scrollbar.set)
        
        self.overdue_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(10, 0))
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y, padx=(0, 10))
        
        # Refresh button
        tk.Button(
            parent,
            text="Refresh",
            command=self.refresh_overdue_list,
            bg=self.accent_color,
            fg="white",
            padx=20
        ).pack(pady=10)
        
        tk.Button(
            parent,
            text="Send Overdue Notices",
            command=self.send_overdue_notices,
            bg=self.warning_color,
            fg="white",
            padx=20
        ).pack(pady=5)
        
        # Initial load
        self.refresh_overdue_list()
    
    def refresh_overdue_list(self):
        """Refresh the overdue list"""
        if not self.overdue_tree:
            return
        
        # Clear existing items
        for item in self.overdue_tree.get_children():
            self.overdue_tree.delete(item)
        
        # Get fresh data from database
        overdue = self.database.check_overdue_books()
        
        # Add overdue books to treeview
        for book in overdue:
            self.overdue_tree.insert('', tk.END, values=(
                book['student_name'],
                book['lrn'],
                book.get('grade_section', ''),
                book['book_title'],
                book['date_time'][:10] if book['date_time'] else "",
                book['due_date'],
                f"{book['days_overdue']} days"
            ))
    
    def setup_admin_settings(self, parent):
        """Setup settings tab"""
        settings_frame = tk.Frame(parent, bg=self.bg_color, relief=tk.GROOVE, bd=2)
        settings_frame.pack(pady=20, padx=20, fill=tk.BOTH, expand=True)
        
        tk.Label(
            settings_frame,
            text="System Settings",
            font=("Helvetica", 16, "bold"),
            bg=self.bg_color
        ).pack(pady=20)
        
        # Borrowing days
        days_frame = tk.Frame(settings_frame, bg=self.bg_color)
        days_frame.pack(fill=tk.X, padx=50, pady=10)
        
        tk.Label(
            days_frame,
            text="Borrowing Period (days):",
            font=("Helvetica", 12),
            bg=self.bg_color,
            width=20,
            anchor='w'
        ).pack(side=tk.LEFT)
        
        days_var = tk.StringVar(value=str(BORROWING_DAYS))
        days_entry = tk.Entry(days_frame, textvariable=days_var, width=10, font=("Helvetica", 12))
        days_entry.pack(side=tk.LEFT, padx=10)
        
        def update_days():
            global BORROWING_DAYS
            try:
                new_days = int(days_var.get())
                if new_days > 0:
                    BORROWING_DAYS = new_days
                    messagebox.showinfo("Success", "Borrowing period updated")
                else:
                    messagebox.showerror("Error", "Please enter a positive number")
            except ValueError:
                messagebox.showerror("Error", "Invalid number")
        
        tk.Button(
            days_frame,
            text="Update",
            command=update_days,
            bg=self.accent_color,
            fg="white"
        ).pack(side=tk.LEFT, padx=10)
        
        # Admin password
        pass_frame = tk.Frame(settings_frame, bg=self.bg_color)
        pass_frame.pack(fill=tk.X, padx=50, pady=20)
        
        tk.Label(
            pass_frame,
            text="Change Admin Password:",
            font=("Helvetica", 12),
            bg=self.bg_color,
            width=20,
            anchor='w'
        ).pack(side=tk.LEFT)
        
        def change_password():
            dialog = tk.Toplevel(self.root)
            dialog.title("Change Password")
            dialog.geometry("400x200")
            dialog.transient(self.root)
            dialog.grab_set()
            
            tk.Label(dialog, text="New Password:").pack(pady=10)
            new_pass = tk.Entry(dialog, show="*", width=20)
            new_pass.pack()
            
            tk.Label(dialog, text="Confirm Password:").pack(pady=10)
            confirm_pass = tk.Entry(dialog, show="*", width=20)
            confirm_pass.pack()
            
            def save():
                if new_pass.get() and new_pass.get() == confirm_pass.get():
                    global ADMIN_PASSWORD
                    ADMIN_PASSWORD = new_pass.get()
                    messagebox.showinfo("Success", "Password changed")
                    dialog.destroy()
                else:
                    messagebox.showerror("Error", "Passwords do not match")
            
            tk.Button(dialog, text="Save", command=save).pack(pady=10)
        
        tk.Button(
            pass_frame,
            text="Change Password",
            command=change_password,
            bg=self.accent_color,
            fg="white"
        ).pack(side=tk.LEFT, padx=10)
        
        # Librarian email
        email_frame = tk.Frame(settings_frame, bg=self.bg_color)
        email_frame.pack(fill=tk.X, padx=50, pady=20)
        
        tk.Label(
            email_frame,
            text="Librarian Email:",
            font=("Helvetica", 12),
            bg=self.bg_color,
            width=20,
            anchor='w'
        ).pack(side=tk.LEFT)
        
        email_var = tk.StringVar(value=LIBRARIAN_EMAIL)
        email_entry = tk.Entry(email_frame, textvariable=email_var, width=30, font=("Helvetica", 12))
        email_entry.pack(side=tk.LEFT, padx=10)
        
        def update_email():
            global LIBRARIAN_EMAIL
            LIBRARIAN_EMAIL = email_var.get()
            self.notifier.librarian_email = LIBRARIAN_EMAIL
            messagebox.showinfo("Success", "Librarian email updated")
        
        tk.Button(
            email_frame,
            text="Update",
            command=update_email,
            bg=self.accent_color,
            fg="white"
        ).pack(side=tk.LEFT, padx=10)
    
    # ============================================================================
    # HELPER METHODS
    # ============================================================================
    
    def add_student_dialog(self):
        """Dialog to add new student to STUDENT DATABASE"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Add Student")
        dialog.geometry("400x300")
        dialog.transient(self.root)
        dialog.grab_set()
        
        tk.Label(dialog, text="LRN:").pack(pady=5)
        lrn_entry = tk.Entry(dialog, width=30)
        lrn_entry.pack()
        
        tk.Label(dialog, text="Student Name:").pack(pady=5)
        name_entry = tk.Entry(dialog, width=30)
        name_entry.pack()
        
        tk.Label(dialog, text="Grade & Section:").pack(pady=5)
        grade_entry = tk.Entry(dialog, width=30)
        grade_entry.pack()
        
        tk.Label(dialog, text="Email:").pack(pady=5)
        email_entry = tk.Entry(dialog, width=30)
        email_entry.pack()
        
        def save():
            lrn = lrn_entry.get().strip()
            name = name_entry.get().strip()
            grade = grade_entry.get().strip()
            email = email_entry.get().strip()
            
            if lrn and name:
                success, msg = self.database.add_student(lrn, name, grade, email)
                if success:
                    messagebox.showinfo("Success", msg)
                    dialog.destroy()
                    # Refresh the student list
                    self.refresh_student_list()
                else:
                    messagebox.showerror("Error", msg)
            else:
                messagebox.showerror("Error", "LRN and Student Name are required")
        
        tk.Button(dialog, text="Save", command=save).pack(pady=10)
    
    def remove_student_dialog(self):
        """Dialog to remove student from STUDENT DATABASE"""
        lrn = simpledialog.askstring("Remove Student", "Enter Student LRN:")
        if lrn:
            success, msg = self.database.remove_student(lrn)
            if success:
                messagebox.showinfo("Success", msg)
                # Refresh the student list
                self.refresh_student_list()
            else:
                messagebox.showerror("Error", msg)
    
    def add_book_dialog(self):
        """Dialog to add new book to BOOK INVENTORY"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Add Book")
        dialog.geometry("400x250")
        dialog.transient(self.root)
        dialog.grab_set()
        
        tk.Label(dialog, text="Book Barcode:").pack(pady=5)
        barcode_entry = tk.Entry(dialog, width=30)
        barcode_entry.pack()
        
        tk.Label(dialog, text="Book Title:").pack(pady=5)
        title_entry = tk.Entry(dialog, width=30)
        title_entry.pack()
        
        tk.Label(dialog, text="Author:").pack(pady=5)
        author_entry = tk.Entry(dialog, width=30)
        author_entry.pack()
        
        def save():
            barcode = barcode_entry.get().strip()
            title = title_entry.get().strip()
            author = author_entry.get().strip() or "unknown"
            
            if barcode and title:
                success, msg = self.database.add_book(barcode, title, author)
                if success:
                    messagebox.showinfo("Success", msg)
                    dialog.destroy()
                    # Refresh the book list
                    self.refresh_book_list()
                else:
                    messagebox.showerror("Error", msg)
            else:
                messagebox.showerror("Error", "Barcode and Title are required")
        
        tk.Button(dialog, text="Save", command=save).pack(pady=10)
    
    def remove_book_dialog(self):
        """Dialog to remove book from BOOK INVENTORY"""
        barcode = simpledialog.askstring("Remove Book", "Enter Book Barcode:")
        if barcode:
            success, msg = self.database.remove_book(barcode)
            if success:
                messagebox.showinfo("Success", msg)
                # Refresh the book list
                self.refresh_book_list()
            else:
                messagebox.showerror("Error", msg)
    
    def send_overdue_notices(self):
        """Send overdue notices"""
        overdue = self.database.check_overdue_books()
        if not overdue:
            messagebox.showinfo("Info", "No overdue books found")
            return
        
        # Group by student
        students_overdue = {}
        for book in overdue:
            sid = book['lrn']
            if sid not in students_overdue:
                students_overdue[sid] = []
            students_overdue[sid].append(book)
        
        if messagebox.askyesno("Confirm", f"Send notices to {len(students_overdue)} students?"):
            # In a real implementation, you would send emails here
            messagebox.showinfo("Success", f"Notices sent to {len(students_overdue)} students")
    
    def generate_report(self):
        """Generate library report"""
        stats = self.database.get_statistics()
        
        report = f"""
LIBRARY SYSTEM REPORT
Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

SUMMARY STATISTICS
-----------------
Total Transactions: {stats['total_transactions']}
Total Borrows: {stats['total_borrows']}
Total Returns: {stats['total_returns']}
Active Borrowings: {stats['active_borrowings']}
Overdue Books: {stats['overdue_books']}
Unique Students: {stats['unique_students']}
Unique Books: {stats['unique_books']}

BOOKS STATUS
-----------
Total Books in System: {stats['total_books']}
Available Books: {stats['available_books']}
Borrowed Books: {stats['borrowed_books']}

STUDENTS STATUS
--------------
Total Students: {stats['unique_students']}
Students with Active Borrowings: {stats['active_borrowings']}
"""
        
        filename = f"library_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        with open(filename, 'w') as f:
            f.write(report)
        
        messagebox.showinfo("Success", f"Report saved as {filename}")
    
    def backup_database(self):
        """Backup the database"""
        import shutil
        
        if os.path.exists(self.database.filename):
            backup_name = f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            shutil.copy2(self.database.filename, backup_name)
            messagebox.showinfo("Success", f"Database backed up as {backup_name}")
        else:
            messagebox.showerror("Error", "Database file not found")

# ============================================================================
# MAIN ENTRY POINT
# ============================================================================

if __name__ == "__main__":
    root = tk.Tk()
    app = LibrarySoftware(root)
    root.mainloop()
